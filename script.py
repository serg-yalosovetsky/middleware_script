import argparse
from json.decoder import JSONDecodeError
from typing import Dict
import aiohttp
import xlsxwriter
import json, yaml
import requests, asyncio
from openpyxl import load_workbook
import os

           
def ifIter(v):
    dictionary = {1:1}
    lists = [0]
    if type(v) == type(dictionary) or type(v) == type(lists):
        return 1
    else:
        return 0
    
    
def ifList(v):
    lists = [0]
    if type(v) == type(lists) and len(v)>0 :
        return 1
    else:
        return 0
  
        
def ifDict(v):
    dictionary = {1:1}
    if type(v) == type(dictionary) and len(v)>0:
        return 1
    else:
        return 0
    
    
def printDict(v0):
    if ifDict(v0):
        for k1,v1 in v0.items():
            return(k1, v1)
    elif ifList(v0):
        for v1 in v0:
            return(v1)
    else:
        return(v0)
         
         
def random_chr(n,m=0,k=1, binary=1):
    if binary:
        rand=''
        if m==-1:
            n,m=m,n
        for i in range(k):
            rand+= chr(n+i)
        return rand
    return chr(n)
    
    
def url_creator(point, url_d={}, debug=0):
    if url_d!={}:
        
                    
        pass
    else:
        
        mw= "mw-tst.itsmartflex.com/" 
        #mw = "mw-apitst.vodafone.ua/" #mw_v
        
        phone_rp = "380662584192"
        phone_сs= "380662583046"
        phone_b= "380662583045"
        phone_o= "380662584418"
        product_id = "?product.id="
        siebel_token = "?siebelToken=177C50397B39937400431DC33E982F3D"
        
        point_s = "MYVF-SETTINGS"
        point_l = "MYVF-LINKS"
                
        get_token = "uaa/oauth/token?grant_type=client_credentials"
        relat_phone = "customer/api/customerManagement/v3/customer/"
        counters_i_bonus = "prepaybalance/tmf-api/prepayBalanceManagement/v4/bucket"
        offer = "qualification/api/productOfferingQualificationManagement/v1/productOfferingQualification/"
        setting_i_links = "entity/api/functions/"
        
        proto = "https"
        slash = "://"
        
    if (point=="token"):
        url = proto + slash + mw + get_token
    if (point=="relPhone"):
        url = proto + slash + mw + relat_phone + phone_rp
    if (point=="countMain" or point=="countDpi"):
        url = proto + slash + mw + counters_i_bonus + product_id + phone_сs
    if (point=="bonus"):
        url = proto + slash + mw + counters_i_bonus + '/' + phone_b
    if (point=="offer"):
        url = proto + slash + mw + offer + phone_o + siebel_token
    if (point=="settings"):
        url = proto + slash + mw + setting_i_links + point_s
    if (point=="links"):
        url = proto + slash + mw + setting_i_links + point_l
        
    log_print(debug, 2, 'url_creator ',url, point)
    
    return url


def header_creator(point, token='', debug=0, fil_type = 0, filler= ''):
 
    log_print(debug, 2, 'heador_creator', filler)
    filler0 = filler if fil_type>=0 else ''
    filler1 = filler if fil_type>=1 else ''
    filler2 = filler if fil_type>=2 else ''
    filler3 = filler if fil_type>=3 else ''
    filler4 = filler if fil_type>=4 else ''
    filler5 = filler if fil_type>=5 else ''
    filler6 = filler if fil_type>=6 else ''
    filler7 = filler if fil_type>=7 else ''
    lang = ['Accept-Language'+filler6 , 'ru'+filler0]
    cont = ['Content-Type'+filler5 , "application/json"+filler1]
    auth = ['Authorization'+filler4 , 'Basic aW50ZXJuYWw6aW50ZXJuYWw='+filler2 ]
    auth2 = ['Authorization'+filler4 , 'Bearer ' + token +filler2 ]
    prof = ['Profile' + filler7] 
    if (point=="token"):
        headers = {lang[0] : lang[1],
                   cont[0] : cont[1],
                   auth[0] : auth[1]  }
    if (point=="settings" or point=="links"):
        headers = {lang[0] : lang[1],
                   cont[0] : cont[1],
                   auth2[0] : auth2[1] }
    if (point=="relPhone"):
        headers = {prof[0] : 'RELATED-PARTY' + filler3,
                   lang[0] : lang[1],
                   cont[0] : cont[1],
                   auth2[0] : auth2[1] }
    if (point=="countMain"):
        headers = {prof[0]: 'COUNTERS' + filler3,
                   lang[0] : lang[1],
                   cont[0] : cont[1],
                   auth2[0] : auth2[1] }
    if (point=="countDpi"):
        headers = {prof[0]: 'COUNTERS-DATA' + filler3,
                   lang[0] : lang[1],
                   cont[0] : cont[1],
                   auth2[0] : auth2[1] }
    if (point=="bonus"):
        headers = {prof[0]: 'BONUS' + filler3,
                   lang[0] : lang[1],
                   cont[0] : cont[1],
                   auth2[0] : auth2[1] }
    if (point=="offer"):
        headers = {prof[0]: 'ACTIVE-OFFER' + filler3,
                   lang[0] : lang[1],
                   cont[0] : cont[1],
                   auth2[0] : auth2[1] }
    
    log_print(debug, 4,'header_creator2',point)
    log_print(debug, 5,'header_creator3',headers)
    return headers
         
              
def data_creator(s, data = {}):
    if data:    
        for d in data:
            data = {d +  s : data[d] + s} 
    else:
            data = {s : s} 
    return data
 
 
def manageRequest(point, verb , token=None, headers={}, data={}, debug=0, log={} ):
    if token==None:
        url = url_creator('token', {}, debug)
        headers = header_creator('token', '', debug)
        r = requests.request(method = 'post', url= url, headers=headers)
        # r = _requests(url, 'post', headers, data, debug)
        token = r.json()["access_token"]
        log_print(debug, 4,'manageRequest token', token)
    
    url = url_creator(point, {}, debug)
    headers = header_creator(point, token, debug)
    r = requests.request(method = verb, url= url, headers=headers)
    # r = _requests(url, verb, headers, data, debug)    
    return (r, verb, url, point, r.json())    


def many_req_gen(verbs, points, token, data={}, debug=0 ):
    
    for point in points:
        log_print(debug, 1, 'many_rew_gen0', point, 'gen' )
        for verb in verbs:
            log_print(debug, 2, 'many_rew_gen1', verb, 'gen' )
            url = url_creator(point, {}, debug)
            headers = header_creator(point, token, debug)        
            x = (verb, url, point, headers, data, debug)
            log_print(debug, 3,'many_rew_gen2', x)
            log_print(debug, 4,'many_rew_gen3', 'yield')
            yield x


def create_filler(n, s='a'):
    st = ''
    for i in range(n):
        st+= s
    return st


def log_print(debug, n, *message):
    if debug>= n:
        print(*message)

 
def gen_u(point, quirks, debug=0):
    if quirks['change'] == 'url':
        if quirks['mode'] == 'content':
            filler = quirks['chr'] * quirks['content_num']
            for i in range(quirks['step']):
                url = url_creator(point, {}, debug) + filler
                filler = filler + quirks['chr']* quirks['content_num']
                log_print(debug, 4, 'gen_u', url)
                yield url
                
        elif quirks['mode'] == 'change':
            for i in quirks['stop'] - quirks['start']:
                filler =  chr( [quirks['start'] +i ] ) 
                url = url_creator(point, {}, debug) + filler
                log_print(debug, 4, 'gen_u', url)
                yield url
    else:
        while True:
            yield url_creator(point, {}, debug)
 
 
def gen_h(point, token, quirks, debug=0):
    log_print(debug, 4, 'gen_h 0', quirks)
    if quirks['change'] == 'header':
        log_print(debug, 4, 'gen_h 1')
        if quirks['mode'] == 'content':
            log_print(debug, 4, 'gen_h 2')
            filler = quirks['chr'] * quirks['content_num']
            for i in range(quirks['step']):
                log_print(debug, 4, 'gen_h 3')
                header = header_creator(point, token, debug, quirks['change_type'], filler)
                log_print(debug, 4, 'gen_h3.5 lenof filler', len(filler))
                filler = filler + quirks['chr']* quirks['content_num']
                log_print(debug, 4, 'gen_h 4',header)
                yield header
                
        elif quirks['mode'] == 'change':
            for i in quirks['stop'] - quirks['start']:
                filler =  chr( [quirks['start'] +i ] ) 
                header = header_creator(point, token, debug, quirks['change_type'], filler)
                log_print(debug, 4, 'gen_h6', header)
                yield header
    else:
        while True:
            yield header_creator(point, token, debug)


def gen_d(quirks, data={}, debug=0):
    if quirks['change'] == 'data':
        log_print(debug, 4, 'gen_d 0')
        if quirks['mode'] == 'content':
            log_print(debug, 4, 'gen_d 1')
            filler = quirks['chr'] * quirks['content_num']
            for i in range(quirks['step']):
                log_print(debug, 4, 'gen_d 2')
                data = data_creator(filler, data) 
                filler = filler + quirks['chr']* quirks['content_num']
                log_print(debug, 4, 'gen_d 4', data)
                yield data
                
        elif quirks['mode'] == 'change':
            for i in quirks['stop'] - quirks['start']:
                data = data_creator(filler, data) 
                filler =  chr( [quirks['start'] +i ] ) 
                log_print(debug, 4, 'gen_d', data)
                yield data
    else:
        while True:
            yield data
     
           
def many_req_quirks_gen(verbs, points, token, quirks, data, debug ):
    
    log_print(debug, 4,' many_req_quirks_gen quirks mode')
    for point in points:
        for verb in verbs:
            subgen_h = gen_h(point, token, quirks, debug)
            subgen_u = gen_u(point, quirks, debug)
            subgen_d = gen_d(quirks, data, debug)
            if quirks['mode'] == 'change':
                n = (quirks['stop_chr'] - quirks['start_chr'] )/ quirks['step_chr']
            if quirks['mode'] == 'content':
                n = quirks['step']
                 
            for i in range(n):
                header = next(subgen_h)
                url = next(subgen_u)
                data = next(subgen_d)
                x = (verb, url, point, header, data, i)

                log_print(debug, 4, ' many_req_quirks_gen', i,n)
            yield x                       
   
       
async def _requests(verbs, points, token, quirks=None, headers = {}, data= {}, debug=0):
    if quirks is not None:
        gen = many_req_quirks_gen(verbs, points, token, quirks, data, debug )
        log_print(debug,2,'_requests1')
    else:
        gen = many_req_gen(verbs, points, token, data, debug )
        log_print(debug,2,'_requests2')
    log_print(debug,2,'_requests3')
    async with aiohttp.ClientSession() as session:
        log_print(debug,2,'_requests session')
        q=0
        try:
            while x:= next(gen):
                q+=1
                verb, url, point, headers, data, filler = x #next(gen)
                async with session.request(verb, url, headers=headers, data=data) as response:
                # async with session.post(url, headers=headers) as response:
                    status =  response.status
                    log_print(debug, 1, '_requests, in session', status)
                    try:
                        resp = await response.json()
                    except aiohttp.client_exceptions.ContentTypeError:
                        log_print(debug, 1, '_requests not a json, q', q)
                        resp = await response.text()

                    if status>=200 and status<300:
                        log[0][(point, url, verb, filler)] = (status, headers, data, resp)
                    else:
                        log[1][(point, url, verb, filler)] = (status, headers, data, resp)
            await session.close()
        except StopIteration as e:
            log_print(debug, 1,'_requests', e, q,filler)


def fetchToken(url=None, headers=None, mw = None):
    if mw is None:
        mw = 'mw-tst.itsmartflex.com'
    elif mw == 'vodafone':
        mw = 'mw-apitst.vodafone.ua' #mw-tst.itsmartflex.com
    
    if url is None:
        url = 'https://' + mw + '/uaa/oauth/token?grant_type=client_credentials'
        url2 = 'https://' + mw + '/uaa/oauth/token?grant_type=password'
    if headers is None:
        basic1 = 'aW50ZXJuYWw6aW50ZXJuYWw='
        basic2= 'c3ZjLXNpZWJlbC1tdzoxMjM0='
        headers = {"Authorization": "Basic " + basic1,"Content-Type": "application/json"}
        headers2 = {"Authorization": "Basic c3ZjLXNpZWJlbC1tdzoxMjM0=","Content-Type": "application/x-www-form-urlencode"}
    # data = {'username' :'380952240016'}
    print(url)
    print(headers)
    r = requests.post(url, headers = headers)#, data = data) 
    print(r.status_code)
    json_response = r.json()
    return json_response['access_token']


def fetchPass(token = None, url=None, headers=None, mw = None):
    if mw is None:
        # mw = 'mw-apitst.vodafone.ua'
        mw = 'mw-tst.itsmartflex.com'
    if url is None:
        # url = 'https://' + mw + '/uaa/oauth/token?grant_type=client_credentials'
        url = 'https://' + mw + '/uaa/oauth/token?grant_type=password'
    if token is None:
        token = 'eyJhbGciOiJSUzI1NiIsInR5cCI6IkpXVCJ9.eyJzY29wZSI6WyJtc191YWFfY3JlZGVudGlhbF90eXBlX3Bhc3N3b3JkX2p3dCIsIm9wZW5pZCJdLCJleHAiOjE2MTE3NjI5NzksImFkZGl0aW9uYWxEZXRhaWxzIjp7Im1zaXNkbiI6IjM4MDk1MjI0MDAxNiJ9LCJhdXRob3JpdGllcyI6WyJST0xFX1NZU1RFTSJdLCJqdGkiOiI3NmRiNTEyMS0zZjJmLTQzZTMtYTJjZi1lMTA4NGNjM2YzZTciLCJ0ZW5hbnQiOiJYTSIsImNsaWVudF9pZCI6InN2Yy1zaWViZWwtbXcifQ.d50VVTF50QF-nqvxrmzcXWg9Ih0qRdTlPYdfhi_RKzgDL_RjDtu_sIiv6FWLKGyAgYcpVD-ukL2F8BU53OGAqgEbJo0zJLzVXEIUDcdc-zxZufPf-x6MtFhwSo4Xk6v22esSgPtU-LD2Kk8wfr4_yA0BChu4OXS2uXwvy7qaQS7h889kIyO2pSeHBJdtS0-EjrP4cFebd32hFZ9J_25ockKjJ17y08ubYQh_YMrbOzHlWrfmbYhMVN71uadgOGfRnwkSWzVJ-GThcFasXTuK7ONKa748sCQWG_q4R7QEzZbvDnFk2Dw-VXyb73kOJQKOJY3QblDTVmkqF-yg8DMKGQ'

    if headers is None:
        # headers = {"Authorization": "Basic aW50ZXJuYWw6aW50ZXJuYWw=","Content-Type": "application/json"}
        # –header ‘Content-Type: application/x-www-form-urlencoded’
        headers = {"Authorization": "Basic YXBwLW15dm9kYWZvbmUtbXc6UGg1ZDg2QnpCNg==","Content-Type": "application/x-www-form-urlencoded", 'Profile': 'MYVODAFONE'}
        headers2 = {"Authorization": "Basic c3ZjLXNpZWJlbC1tdzoxMjM0’=","Content-Type": "application/x-www-form-urlencode"}
    data = {'username' :'380952240016', 'password' : token}

    print(url)
    print(headers)
    r = requests.post(url, headers = headers, data = data) 
    print(r.status_code)
    try:
        json_response = r.json()['access_token']
    except:
        json_response = r.text
        print(json_response)
    return json_response


def print_elem(dic, key):
    if type(dic) == type(dictionary) and dic.get(key, 0):
        return key
    else:
        return key + ' ' + dic[key]


def isType(e):  
    '''-1 None, 0 if empty array, 1 if str int float, 2 if list, 3 if dict
    '''
    if e is None:
        return -1
    if e == [] or e == {} or e == ():
        return 0
    if str(type(e)) == "<class 'bool'>" or str(type(e)) == "<class 'int'>" or str(type(e)) == "<class 'float'>" or str(type(e)) == "<class 'str'>":
        return 1 
    if str(type(e)) == "<class 'list'>":
        return 2 
    if str(type(e)) == "<class 'dict'>":
        return 3  


def is_int(i):
    try:
        int(i)
        return 1 
    except:
        return 0
    
    
def none_safe_str(s):
    if s is None:
        return s
    else:
        return str(s)
    
    
def checkArray(el): #0 str, 1 list, 2 dict, 3 listdict, 4 listlist, 5 dictlist, 6 dictdict
    if isType(el)<2: #isType(): -1 None, 0 if empty array, 1 if str int float, 2 if list, 3 if dict
        return 0
    for e in el:
        # print('checkarray0',isType(e),type(e))
        # print(e)
        if isType(el)==2: #list
            if isType(e)>=2: #list or dict
                return isType(e) + 1
            else: return 1
        if isType(el)==3: #dict
            if isType(el[e])>=2: #list or dict
                return isType(el[e]) + 3
            else: return 2
    return 0        
        
        
def parse_string(s):
    for c in s:
        if c=='[':
            pass
        if c=='{':
            print()
        # if 


def parse_str(s,n=0, dic=0, dense=0):
    if isType(s)<2:
        yield (s,n,dic)
    if isType(s)==2:
        for c in s:
            yield from parse_str(c, n+1, dic, dense=dense)
        print()
    if isType(s)==3:
        for c in s:
            yield (c,n+1,dic)
            yield from parse_str(s[c], n+dense, dic+1, dense=dense)
        dic -=1
    


def check_if_filed_in_excluding(parent, field, point, settings, reverse=0):

    for p in settings['field of response']:
        if point == p:
            for m in settings['field of response']:
                # if m == field or m in parent:
                if m == field :
                    print(1)
                    return 1
                elif m in parent:
                    print(2)
                    return 1
                else:
                    print(0)
                    return 0


def border_drawer(workbook):
    '''return border, border_no_top, border_no_topright, border_no_left, border_no_right, border_no_bottom, border_no_bottomleft
    '''
    border = workbook.add_format()
    border.set_border(style=1)
    border_no_top = workbook.add_format()
    border_no_top.set_bottom()
    border_no_top.set_left()
    border_no_top.set_right()
    border_no_topright = workbook.add_format()
    border_no_topright.set_border(style =3)
    border_no_topright.set_bottom()
    border_no_topright.set_left()
    border_no_left = workbook.add_format()
    border_no_left.set_border(style =3)
    border_no_left.set_bottom()
    border_no_left.set_top()
    border_no_left.set_right()
    border_no_right = workbook.add_format()
    border_no_right.set_border(style =3)
    border_no_right.set_bottom()
    border_no_right.set_top()
    border_no_right.set_left()
    border_no_bottom = workbook.add_format()
    border_no_bottom.set_top()
    border_no_bottom.set_left()
    border_no_bottom.set_right()
    border_no_bottomleft = workbook.add_format()
    border_no_bottomleft.set_border(style =3)
    border_no_bottomleft.set_top()
    border_no_bottomleft.set_right()
    bd = {'border':border, 'border_no_top':border_no_top, 'border_no_topright':border_no_topright, 
        'border_no_left':border_no_left, 'border_no_right':border_no_right, 'border_no_bottom':border_no_bottom,
         'border_no_bottomleft':border_no_bottomleft}
    return bd


def check_if_cell_is_not_empty(s):
    try:
        if s is not None and s != '':
            return 1
        else:
            return 0
    except Exception as e:
        print(e)
       

def append_list_in_dict(dict_, list_, elem):
    print(dict_, list_, elem)
    try:
        dict_[list_].append(elem)
    except:
        dict_[list_] = []
        dict_[list_].append(elem)
    return  dict_


def add_record_in_dict(dict_, key, key2, elem):
    print(dict_, key, key2, elem)
    try:
        dict_[key][key2] = elem
    except:
        dict_[key] = {}
        dict_[key][key2] = elem
    return  dict_
      
      
       
def parse_settings_in_read_settings(settings, c, mode):
    if c[0].value == 'начальная точка': settings['diapasone'].append(c[1].value)
    if c[0].value == 'конечная точка': settings['diapasone'].append(c[1].value)
    if c[0].value == 'уровень раскрытия ответа запроса': settings['verbose_level'] = int(c[1].value)   
    if c[0].value == 'url для токена миддлваре': url_for_token_mv = c[1].value
    if c[0].value == 'генерация токена миддлваре': gen_for_token_mv = c[1].value
    if c[0].value == 'url для токена сибеля':  url_for_token_siebel = c[1].value            
    if c[0].value == 'генерация токена сибеля':  gen_for_token_siebel = c[1].value            
    if c[0].value == 'проверка определения способа подстановки токена': url_for_token_check = c[1].value
    if c[0].value == 'токен миддлваре': 
        if c[1].value == 'из точки':
            settings['url_mw']['from'] = 'from_point_and_url'            
            settings['url_mw']['from_point'] = c[2].value            
            settings['url_mw']['from_url'] = url_for_token_mv           
            settings['url_mw']['how_gen'] = gen_for_token_mv     
            settings['url_mw']['how_check_url'] = url_for_token_check     
        if c[1].value == 'взять из клетки':
            settings['url_mw']['from'] = 'from_cell'            
            settings['url_mw']['token'] = c[2].value            
        if c[1].value == 'ничего не делать':
            settings['url_mw']['from'] = 'nothing'            
    if c[0].value == 'токен сибель': 
        if c[1].value == 'из точки':
            settings['url_siebel']['from'] = 'from_point_and_url'            
            settings['url_siebel']['from_point'] = c[2].value            
            settings['url_siebel']['from_url'] = url_for_token_siebel           
            settings['url_siebel']['how_gen'] = gen_for_token_siebel           
            settings['url_siebel']['how_check_url'] = url_for_token_check     
        if c[1].value == 'взять из клетки':
            settings['url_siebel']['from'] = 'from_cell'            
            settings['url_siebel']['token'] = c[2].value            
        if c[1].value == 'ничего не делать':
            settings['url_siebel']['from'] = 'nothing'            

    if c[0].value == 'старый токен': settings['verbose_level'].append(c[1].value)            
    if c[0].value == 'проверка совпадения ответа': 
        if c[1].value == 'простая': 
            settings['how to check response'] ='simple'            
        if c[1].value == 'сложная': 
            settings['how to check response'] = 'complicated'            
            if c[2].value == 'исключить': 
                settings['exclude or include fields'] = 'exclude'            
                mode = 'mode_exc_or_inc_values_in_cheking_response'        
            if c[2].value == 'включить': 
                settings['exclude or include fields'] = 'include'    
                mode = 'mode_exc_or_inc_values_in_cheking_response'        
    if c[0].value == 'вывод типов полей в ответе': settings['show type of field in response'] = c[1].value            
    if c[0].value == 'за каждым элементом списка ответов перевод строки ': settings['enter after each element of list'] = c[1].value            
    if c[0].value == 'ответ в рамочки': settings['response in borders'] = c[1].value            
    if c[0].value == 'размер столбцов': 
        mode = 'mode_parse_columns_width'
        settings['columns'][c[1].value] = c[2].value            
    if c[0].value == 'получение значений' : 
        if c[1].value == 'да':
            mode= 'mode_parse_for_getter'
            
    if c[0].value == 'подстановка значений' : 
        if c[1].value == 'да':
            mode= 'mode_parse_for_setter'
            
    return settings, mode
        
       
def read_settings(filename, sheet_name='settings', diapasone=('A1', 'C100'), validate_function=None, **kwargs ):
    
    wb = load_workbook(filename = filename)
    if sheet_name == '__active':
        sheet = wb.active
    else: 
        sheet = wb.get_sheet_by_name(sheet_name)
    try:
        cells = sheet[diapasone[0]: diapasone[1]]
    except Exception as e:
        printlog('Неверный диапазон значений для парсинга')
    c = [0 for i in range(len(cells) )]
    parsed_settings = {'diapasone' : [], 'verbose_level' : 0, 'url_mw' : {}, 'url_siebel' : {}, 'how to check response' : '', 'exclude or include fields': '',
                            'field of response': {},'show type of field in response' : '', 'enter after each element of list': '', 'response in borders':'', 
                            'columns' : {}, 'token-mw' :'', 'token-siebel' :'', 'getter':{}, 'setter':{}, 'getter_inv':{}, 'setter_inv':{}}

    mode = ''
    i = 0
    for *c, in cells:  
        print(c[0].value,c[1].value,c[2].value)
                
        if check_if_cell_is_not_empty(c[0].value):

            if not (mode == 'mode_parse_for_setter' or mode == 'mode_parse_for_getter'):
                parsed_setings, mode = parse_settings_in_read_settings(parsed_settings, c, mode)
            else:
                if mode == 'mode_parse_for_getter':
                    # parsed_settings['getter'][c[0].value] = {c[2].value:c[1].value}
                    add_record_in_dict(parsed_settings['getter'], c[0].value, c[2].value, c[1].value)
                    add_record_in_dict(parsed_settings['getter_inv'], c[0].value, c[1].value, c[2].value)
                if mode == 'mode_parse_for_setter':
                    # parsed_settings['setter'][c[0].value] = {c[2].value:c[1].value}
                    add_record_in_dict(parsed_settings['setter'], c[0].value, c[2].value, c[1].value)
                    add_record_in_dict(parsed_settings['setter_inv'], c[0].value, c[1].value, c[2].value)


        else:
            if mode == 'mode_parse_for_setter' or mode == 'mode_parse_for_getter':
                mode = ''
            
            if check_if_cell_is_not_empty(c[1].value):
            
                if mode == 'mode_exc_or_inc_values_in_cheking_response':
                    append_list_in_dict(parsed_settings['field of response'], c[1].value, c[2].value )

                if mode == 'mode_parse_columns_width':
                    parsed_settings['columns'][c[1].value] = c[2].value            
        
    return parsed_settings
        

def parse_2_dict(*args):
    s = ''
    i=1
    if len(args)==1:
        return args[0]
    for a in args:
        if i==1:
            i+=1
            continue
        s += '{"'+a+'":'
        i+=1
    for j in range(i):
        s+='}'
    return s



def json_parse(s):
    if s=='' or s is None:
        return s
    try:
        str_json = json.loads( s.replace("'",'"'))
    except:
        try:
            q0 = s.find('"')
            q1 = s.rfind('"')
            s = s[:q0]+ s[q0:q1].replace("'",  '') +s[q1:]
            s = s.replace("'", '"')
            str_json = yaml.load(s)
        except Exception as e:
            print(e)
            str_json = 'json decode error, maybe unexpectend of string'
    return str_json


def wrapper_for_json_parse(value):
    s = value
    try:
        str_json = json_parse(s)
    except Exception as e:
        print(e)
        str_json = ''
    return str_json
        
        
def parsing_response_and_post_data(parsed_data, data, point, verb, safe=0, type_='parsed_response'):
    str_json = wrapper_for_json_parse(data)
    if safe:
        if str_json == '':
            return parsed_data
    parsed_data[type_][point] = {}
    parsed_data[type_][point][verb] = str_json  
    return parsed_data
        
        
def read_data(filename, settings='', sheet_name='__active', diapasone=('A1', 'I40'), validate_function=None, **kwargs ):
    
    wb = load_workbook(filename = filename)
    if sheet_name == '__active':
        sheet = wb.active
    else: 
        sheet = wb[sheet_name]
    try:
        if settings == '':
            cells = sheet[diapasone[0]: diapasone[1]]
        else:
            cells = sheet[settings['diapasone'][0]: settings['diapasone'][1]]
            
    except Exception as e:
        print('Неверный диапазон значений для парсинга')
    c = [0 for i in range(len(cells) )]
    parsed_data = {'points' : [], 'verbs' : {}, 'headers_name' : {}, 'headers_value' : {}, 'url' : {}, 
                   'post_data':{}, 'parsed_post_data':{}, 'response' : {}, 'parsed_response' : {}, 'response_code' : {}, 'token' :0}

    i = 0
    for *c, in cells:  
        i +=1
        if i==1:
            continue
        if c[1].value is not None and is_int(c[0].value) :
            parsed_data['points'].append(c[1].value)
            last_point = c[1].value
            parsed_data['verbs'][c[1].value] = []
            parsed_data['verbs'][c[1].value].append(c[2].value)
            parsed_data['headers_name'][c[1].value] = []
            parsed_data['headers_name'][c[1].value].append(c[3].value)
            parsed_data['headers_value'][c[1].value] = {}
            parsed_data['headers_value'][c[1].value][c[3].value] = c[4].value
            parsed_data['url'][c[1].value] = []
            parsed_data['url'][c[1].value].append(c[5].value)
            parsed_data['post_data'][c[1].value] = {}
            parsed_data['post_data'][c[1].value][c[2].value] = c[6].value
            parsed_data['parsed_post_data'][c[1].value] = {}
            str_json = wrapper_for_json_parse(c[6].value)
            parsed_data['parsed_post_data'][c[1].value][c[2].value] = str_json                
            parsed_data['response_code'][c[1].value] = {}
            parsed_data['response_code'][c[1].value][c[2].value] = c[7].value
            parsed_data['response'][c[1].value] = {}
            parsed_data['response'][c[1].value][c[2].value] = c[8].value                
            parsed_data['parsed_response'][c[1].value] = {}
            str_json = wrapper_for_json_parse(c[8].value)
            parsed_data['parsed_response'][c[1].value][c[2].value] = str_json                
                
        elif last_point is not None:
            if c[2].value is not None:
                parsed_data['verbs'][last_point].append(c[2].value)
            if c[3].value is not None:
                parsed_data['headers_name'][last_point].append(c[3].value)
                if c[4].value is not None:
                    parsed_data['headers_value'][last_point][c[3].value] = c[4].value
                else:
                    parsed_data['headers_value'][last_point][c[3].value] = ''
            if c[5].value is not None:
                parsed_data['url'][last_point].append(c[5].value)
            if c[6].value is not None:
                parsed_data['post_data'][last_point].append(c[6].value)
            if c[7].value is not None:
                parsed_data['response_code'][last_point][c[2].value] = c[7].value
            if c[8].value is not None:
                parsed_data['response'][last_point][c[2].value] = c[8].value
                str_json = wrapper_for_json_parse(c[8].value)
                parsed_data['parsed_response'][last_point][c[2].value] = str_json  
            
    return parsed_data




def reading_response_from_file(filename, parsed_data, point='', verb='' ):
    
    with open(filename, encoding='utf-8' ) as f:
        for line in f.readlines():
            for lin in line.split('/n'):
                points = []
                i = line.find('{')
                for l in lin[:i].split():
                    points.append(l)
                if point!='' and verb !='':
                    parsed_data['response'][points[0]][points[1]] = lin[i:]
                else:
                    parsed_data['response'][point][verb] = lin[i:]
    return parsed_data
                    
    
def filling_getters(settings, parsed_data, getters ):             
    
    for points in parsed_data['parsed_response']:
        if points in settings['getter'].keys():
            for verbs in parsed_data['parsed_response'][points]:
                if isType(parsed_data['parsed_response'][points][verbs]) >=2:
                    for r in parsed_data['parsed_response'][points][verbs]:
                        if isType(parsed_data['parsed_response'][points][verbs][r]) >=2:
                            for re in parsed_data['parsed_response'][points][verbs][r]:
                                if isType(parsed_data['parsed_response'][points][verbs][r][re]) >=2:
                                    for res in parsed_data['parsed_response'][points][verbs][r][re]:
                                        if isType(parsed_data['parsed_response'][points][verbs][r][re][res]) >=2:
                                            for resp in parsed_data['parsed_response'][points][verbs][r][re][res]:
                                                if isType(parsed_data['parsed_response'][points][verbs][r][re][res][resp]) >=2:
                                                    for respo in parsed_data['parsed_response'][points][verbs][r][re][res][resp]:
                                                        if points == 'loginV2':print(respo)
                                                        if settings['getter'].get(points,0):
                                                            if respo in list(settings['getter_inv'][points].keys()):
                                                                alias = list(settings['getter'][points].keys())[0]
                                                                getters[alias] = parsed_data['parsed_response'][points][verbs][r][re][res][resp][respo]
                                                else:
                                                    if points == 'loginV2':print(resp)
                                                    
                                                    if settings['getter'].get(points,0):    
                                                        if resp in list(settings['getter_inv'][points].keys()):
                                                            alias = list(settings['getter'][points].keys())[0]
                                                            getters[alias] = parsed_data['parsed_response'][points][verbs][r][re][resp]                   
                                        else:
                                            if points == 'loginV2':print(res)
                                            
                                            if settings['getter'].get(points,0):    
                                                if res in list(settings['getter_inv'][points].keys()):
                                                    alias = list(settings['getter'][points].keys())[0]
                                                    getters[alias] = parsed_data['parsed_response'][points][verbs][r][re][res]      
                                else:
                                    if points == 'loginV2':print(re)
                                    
                                    if settings['getter'].get(points,0):
                                        if re in list(settings['getter_inv'][points].keys()):
                                            alias = list(settings['getter'][points].keys())[0]
                                            getters[alias] = parsed_data['parsed_response'][points][verbs][r][re]
 
                        else:
                            if points == 'loginV2':print(r)
                            
                            if settings['getter'].get(points,0):
                                if r in list(settings['getter_inv'][points].keys()):
                                    alias = list(settings['getter'][points].keys())[0]
                                    getters[alias] = parsed_data['parsed_response'][points][verbs][r]
    return getters
 
def filling_setters(settings, setters, getters):
    
    for sett in settings['setter']:
        for set in settings['setter'][sett]: 
            print('set= ', set)
            var = settings['setter'][sett][set]
            print('var= ', var)
            i0 = set.find('{{')
            i1 = set.find('}}')
            print('io i1 ', i0, i1)
            if i0>=0 and i1>=0 :
                print('set[i0+2:i1]', set[i0+2:i1])
                print(set[:i0], getters[set[i0+2:i1]], set[i1+2:], sep='')
                setters[set] = set[:i0] + getters[set[i0+2:i1]] + set[i1+2:]
            else:
                setters[set] = set
    return setters
 

settings = read_settings(filename='setting1.xlsx', sheet_name='settings')        
parsed_data = read_data(filename='setting1.xlsx',settings = settings , sheet_name='data')    
setters = {}
getters = {}
getters = filling_getters(settings, parsed_data, getters )             
setters = filling_setters(settings, setters, getters)



        
parent = ['']
parent.append(1)
import copy
p0 = copy(parent)
parent

def parse_dict(s, list_, parent=[], n=0, dic=0):
    print()
    # print('s,list_', s,list_)
    if s in list_:
        print('||||parseDict', s, parent)
        yield s, parent
    if isType(s)<2:
        # print(s)
        # yield (s,list_, n,dic)
        yield '', parent
    if isType(s)==2:
        for c in s:
            yield from parse_dict(c, list_, parent, n+1, dic)
        # print()
    if isType(s)==3:
        for c in s:
            # yield (c,n+1,dic)
            print('parseDict 0', c,'s[c]',s[c], 'par', parent)
            
            if c in list_:
                print('||||parseDict', c, parent)
                yield c, parent
            if isType( s[c]) >=2:    
                parent0= copy.deepcopy(parent)
                parent0.append(c)
                yield from parse_dict(s[c], list_, parent0, n, dic+1)
        dic -=1
    


def substitution_with_setters(settings, parsed_data, getters, setters, field_for_subst=['post_data', 'headers_value']):             

    for p in settings['setter_inv']:
        print(p)
        key = list(settings['setter_inv'][p].keys())
        lis = list(parsed_data['headers_value'].get(p, 0).keys())
        for k in key:
            print(f'key={key}, {k} in {lis}, {k in lis}')
            if k in lis:
                print()
                parsed_data['headers_value'][p][k] = setters[settings['setter_inv'][p][k]] 
                print()
            print()

    for p in settings['setter_inv']:
        print(p)
        key = list(settings['setter_inv'][p].keys())
        lis = parsed_data['post_data'].get(p, 0).items()
        if lis != 0:
            for li_k, li_v in parsed_data['parsed_post_data'].get(p, 0).items():
                # print('key')
                # print(li_k)
                # print(li_v)
                try:
                    _ = li_v.items()
                    for l_k, l_v in li_v.items():
                        try:
                            _ = l_v.items()
                            
                            
                            for _k, _v in l_v.items():
                                if p=='getRelPhone':
                                    print(f'|             {_v}     |{_k}')
                                
                                print(f'|point {p}, value {_v}, {_k} in {key}, {_k in key}')
                                if _k in key:
                                    print()
                                    print("_k", _k)
                                    # parsed_data['parsed_post_data'][p][_k] = setters[settings['setter_inv'][p][_k]] 
                                    print ("parsed_data['parsed_post_data'][p][li_k][l_k][_k]", parsed_data['parsed_post_data'][p][li_k][l_k][_k]) 
                                    print("setters[settings['setter_inv'][p][_k]]", settings['setter_inv'][p][_k], p, _k) 
                                    print()
                                print()
                        except: pass
                except: pass






type(parsed_data['post_data']['getToken']['post'])
res = parsed_data['response']['getlinks']['post']
res = str(res)
parsed_data['post_data']['getToken'].get('post','')
parsed_data['url']
write_2_excel_parsed_data(parsed_data, settings )

resp
type(res)
dict(res)
import json,yaml
resp =  res.replace("'", "\"")
d= yaml.load(s)
repr(res)

settings

res2 =res
res3 = json.loads(s3.replace("'", '"'))
res3.keys()
s2 = '{"loginV2": {"error": 0, "values": {"tempToken": "8DBA35BBA4C486592B1BCC0A984DC5BE"}},"identification": {"error": 0, "values": { "id":""}}}'
g = parse_str(res3, 0)
s,n,d = next(g)
s
s2
s3 = "{'access_token': 'eyJhbGciOiJSUzI1NiIsInR5cCI6IkpXVCJ9.eyJzY29wZSI6WyJvcGVuaWQiXSwiZXhwIjoxNjE2NDUwMzI4LCJhdXRob3JpdGllcyI6WyJTVVBFUi1BRE1JTiJdLCJqdGkiOiIyZmNhY2MzYy1mM2I0LTQzNmQtYTYyZC01ZmFlZjdmMGY0YjUiLCJ0ZW5hbnQiOiJYTSIsImNsaWVudF9pZCI6ImludGVybmFsIn0.AaK7IxOhYdjTTvDtx2qa1ndyhjhZN11teSECFcobBnQ8fdIBvsOZzOwqbYv5Kw7cPhh8kGxi5OKjFQqYQjiXIeZFpDf6BzKOvwBu2QUskdFxITpnuLxXrwhemXNHbsuLn1M46LOEmRqYJ8nl6QHEZyEB8SePSNZIWBHCUbGJKWIMoMxdB6PjQjnSSw0WbHsYj4petYEVn56pj3sppyxpr3vppjF3HWM7XtoNvzee9P6fLaKtwuEgDfOXbUYGRNiYPttkV5sR7UjWaEOd4zIklsu2vOetrojiteeVtUALCaUqNaYUoczfRMbvHdeucetrD5Lsi8wmN2oTbvM1Falmzg', 'token_type': 'bearer', 'expires_in': 86399, 'scope': 'openid', 'tenant': 'XM', 'jti': '2fcacc3c-f3b4-436d-a62d-5faef7f0f4b5'}"




settings

def excel_with_border(prev, curr, new,  border, end=0):
    '''curr =[x,y,s] prev=[x0,y0,s0]
    '''
    if end==0:
        if curr[0]>prev(0) and curr[1]==prev(1):
            нарисовать предыдущую без нижней границы 
        if curr[0]==prev(0) and curr[1]>prev(1):
            нарисовать текущую без верхней границы 



def worksheet_write_twice_shift(sheet, x,y, s, shift, s_new, parent=[], point='', checking='', prev='', settings='', param='', x0=0, y0=0, type_s=0): #i,d
    if param=='':
        sheet.write(x,y, str(s))
        if s_new!='':
            sheet.write(x,y+shift, str(s_new))
    else:
        sheet.write(x,y, str(s), param)
        if s_new!='':
            sheet.write(x,y+shift, str(s_new), param)

    if x0!=0 or y0!=0:
            
            if not check_if_filed_in_excluding(parent, s, point, settings, reverse=0):
                if s==s_new: check = 1
                else: check = 0
            else:
                check =1
            

            sheet.write(x,y0, str(bool(check)) )
            if type_s: sheet.write(x,y0+1, str(type(s)) )

    else:
        check = 0    
    return check


def _write_post_data_to_excel( data, point='', filename='test.xlsx', verbose=10, workbook='', current_worksheet='',new_worksheet='testings',border_draw=1,x0=0,y0=0):
    '''Функция для красивого вывода ответа на запрос в файл екселя
    '''
    gen = parse_str(data,0)
        
    if workbook=='':
        workbook = xlsxwriter.Workbook(filename)
        need_close = 1
    else:
        need_close =0
        
    if current_worksheet=='':
        worksheet = workbook.add_worksheet(new_worksheet)
    else:
        worksheet = current_worksheet
    if border_draw ==1:
        bd = border_drawer(workbook)

    try:
        i=0
        shift_right_prev = 0
        shift_right_max = 0
        str_resp_old=''
        counter=0
        border_prev = ''
        print('init')
        while True:
            str_resp,shift_down,shift_right =next(gen)
            print(str_resp)
            shift_right+=y0
            if i==0:
                shift_right_prev=shift_right
                n_prev=shift_down
                     
            if counter==0:
                i = x0
                shift_right_prev=shift_right
                i_prev=i
            if shift_right>shift_right_prev:
                i-=1
            checking = ''
            print('postinit')

            if shift_down == n_prev and counter>0: #если уровень списка одинаков для текущего и предыдущего
                if shift_right==shift_right_prev: #если уровень словаря одинаков для текущего и предыдущего
                    if border_prev == 'no left': #если для предыдущей клеточки установлена граница "без левой стороны", то устанавливаем "без нижней и левой"
                        worksheet.write(i_prev, shift_right_prev, str(str_resp_old), bd['border_no_bottomleft'])
                            
                    else: #иначе устанавливаем только  "без нижней"
                        worksheet.write(i_prev, shift_right_prev, str(str_resp_old),bd['border_no_bottom'])

                    # sett.write(i, d, str(s),border_no_top)
                    worksheet.write(i, shift_right,str_resp, bd['border_no_bottom'])
                    border_prev = 'no top' #устанавливаем "без верхней " для текущей
                    counter+=1
                    print('shift_down == n_prev and counter>0')

                elif shift_right>shift_right_prev:
                    if border_prev == 'no top': #если для предыдущей клеточки установлена граница "без правой стороны", то устанавливаем "без верхней и правой"
                        worksheet.write(i_prev, shift_right_prev, str(str_resp_old),bd['border_no_topright'])
                    else: #иначе устанавливаем только "без правой"
                        worksheet.write(i_prev, shift_right_prev, str(str_resp_old),bd['border_no_right'])
                    # sett.write(i, d, str(s),border_no_left)
                    worksheet.write(i, shift_right,str(str_resp),bd['border_no_left'])
                    border_prev = 'no left'
                    counter+=1
                    print('shift_right>shift_right_prev')
                else:
                    # sett.write(i, d, str(s))
                    worksheet.write(i, shift_right,str(str_resp))
                    border_prev = ''
                    counter+=1
                    print('shift_right>shift_right_prev else')

            else:
                # sett.write(i, d, str(s))
                worksheet.write(i, shift_right,str(str_resp))
                border_prev = ''
                counter+=1
                print('else else')

            
            shift_right_prev = shift_right
            shift_right_max = max(shift_right_max, shift_right)
            i_prev = i
            n_prev = shift_down
            str_resp_old = str_resp
            i+=1
            print('post init')

    except Exception as e:
        print(e)
    finally:
        if need_close:
            workbook.close()
    return (workbook, worksheet, i-x0+1, shift_right_max - y0+1)


def _write_response_to_excel(resp_old, resp_new, point='', filename='test.xlsx', verbose=10, workbook='', current_worksheet='',new_worksheet='testings',border_draw=1, shift_on_y=3,x0=0,y0=0):
    '''Функция для красивого вывода ответа на запрос в файл екселя
    '''

    gen_old = parse_str(resp_old,x0,y0)
    if resp_new !='':
        gen_new = parse_str(resp_new,x0,y0)
        
    if workbook=='':
        workbook = xlsxwriter.Workbook(filename)
        need_close = 1
    else:
        need_close =0
    #смещения нового респонса относитлеьно старого   
    shift = 4  #max(shift_right, 1) 
    if current_worksheet=='':
        worksheet = workbook.add_worksheet(new_worksheet)
    else:
        worksheet = current_worksheet
    
    try:
        i=0
        shift_right_prev = 0
        shift_right_max = 0
        shift_down_max = 0
        str_resp_old=''
        counter=0
        while True:
            str_resp,shift_down,shift_right =next(gen_new)
            shift_down_max = max(shift_down_max, shift_down)

            str_resp_old,shift_down_old,shift_right_old =next(gen_old)  #str_resp_prev,n2,d2 =next(gen_example) 
                     
            if counter==0:
                i = 0
                shift_right_prev=shift_right
            if shift_right>shift_right_prev:
                i-=1
            checking = ''
            if settings['how to check response'] == 'simple':
                checking = 'simple'
            if settings['how to check response'] == 'complicated':
                if settings['exclude or include fields'] == 'exclude':
                    checking = 'exclude'
                if settings['exclude or include fields'] == 'include':
                    checking = 'include'
            
            worksheet_write_twice_shift(worksheet, shift_down_max+i, shift_right,str_resp, shift_on_y, str_resp_old, parent, point,checking=checking, settings=settings, x0=i, y0=y0-1) #i,d
            counter+=1
            
            shift_right_prev = shift_right
            shift_right_max = max(shift_right_max, shift_right)
            i+=1
            counter+=1
    except Exception as e:
        print(e)
    finally:
        if need_close:
            workbook.close()
    print(f'shift_down-x0+1 { shift_down_max+i-x0+1}, shift_right_max - y0 {shift_right_max-y0 +1}')

    return (workbook, worksheet,  shift_down_max+i-x0, shift_right_max-y0 )

            
def wrapper(func, list_, ws, x,y,s, *args, **kwargs):
    '''
    lis, return_ = wrapper(worksheet_write_twice_shift, lis, worksheet, shift_down_max+i, shift_right,str_resp, shift_on_y, str_resp_old, parent, point,checking=checking, settings=settings, x0=i, y0=y0-1) 
    '''
    if list_ != []:
        print(list_)
        return_ =func(ws, x,y,s *args,prev=list_[-1] ,**kwargs)
    else:
        print('oops')
    x = kwargs['val']
    list_.append([x,y,s])
    return list_, return_ ### DECOMMENT



def write_2_excel_parsed_data(parsed_data, settings, filename='response.xlsx', sheet_name='response', **kwargs ):
    workbook = xlsxwriter.Workbook(filename = filename)
    worksheet = workbook.add_worksheet(sheet_name)
    if settings is not None and settings.get('columns_width',0):
        for c, w in settings['columns_width'].items():
            worksheet.set_column(int(c), int(c), int(w))
    counter = 1
    point_counter = 1  
    for point in parsed_data['points']:
        worksheet.write( counter, point_counter, str(point))
        verb_counter = counter
        print('log05')
        for verb in parsed_data['verbs'][point]:
            # worksheet.write( counter, point_counter+5, none_safe_str( parsed_data['post_data'][point].get(verb,'')) )
            y_data_counter =0
            data = parsed_data['parsed_post_data'][point].get(verb,'')
            print('data')
            print(data)
            print('counter, point_couter', counter, point_counter+5)
            print()
            # post_counter=0
            workbook,worksheet,post_counter, y_data_counter = _write_post_data_to_excel( data, verbose=3, current_worksheet=worksheet, workbook=workbook,x0=counter,y0=point_counter+5)
            print('log0')
            worksheet.write( verb_counter, point_counter+1, str(verb))
            try:
                if verb=='post' and parsed_data['post_data'][point].get(verb,''):
                    r = requests.request(verb, *parsed_data['url'][point], headers=parsed_data['headers_value'][point], data=parsed_data['post_data'][point][verb])
                else:
                    r = requests.request(verb, *parsed_data['url'][point], headers=parsed_data['headers_value'][point])
                resp = r.json()
            except Exception as e:
                print(e)
                resp = r.text
            print('log1')
            resp_counter = 0
            worksheet.write( verb_counter, point_counter+y_data_counter+7, str(parsed_data['response_code'][point][verb])+str(r))
            resp_old = parsed_data['parsed_response'][point][verb]
            print(f'point {point} x0 = verb_counter {verb_counter}, y0= point_counter+y_data_counter+9 {point_counter+y_data_counter+9} type of resp_old {type(resp_old)} {isType(resp_old)}')
            workbook,worksheet,resp_counter, y_data_counter = _write_response_to_excel( resp, resp_old, point = point, shift_on_y=5, verbose=3,
                                                                                       current_worksheet=worksheet, workbook=workbook,x0=verb_counter,y0=point_counter+y_data_counter+9)
            print(f'after resp  {point} x0 = resp_counter {resp_counter}, y0= y_data_counter {y_data_counter} смещение {resp_counter}, {y_data_counter-point_counter+9}')
            # workbook,worksheet,resp_counter, y =  write_to_excel2( resp, resp2, shift=4, current_worksheet=worksheet, workbook=workbook,x0=verb_counter,y0=1+point_counter+y_data_counter+8)
            # workbook,worksheet,resp_counter, y =  write_to_excel2(resp, current_worksheet=worksheet, workbook=workbook,x0=verb_counter,y0=1+point_counter+8)
            # workbook,worksheet,resp_counter2, y2 =  write_to_excel2(parsed_data['response'][point][verb],  current_worksheet=worksheet, workbook=workbook,x0=verb_counter,y0=1+point_counter+6+8)
            
            verb_counter += max(resp_counter, post_counter)+1
        worksheet.write( counter, point_counter+2, str(parsed_data['url'][point][0]))
        header_counter = counter
        for h in parsed_data['headers_name'][point]:
            worksheet.write( header_counter, point_counter+3, str(h))
            worksheet.write( header_counter, point_counter+4, str(parsed_data['headers_value'][point][h]))
            header_counter += 1
        counter = max(counter+1, verb_counter, header_counter) + 1
    workbook.close()


    
    
    
    
    
    



def change_dir():
    cwd = os.getcwd()
    cwd0 = os.getcwd()
    os.chdir(r"C:\Users\syalosovetskyi\Downloads\python\middleware_script")
    os.listdir('.')
    cwd

def prompt(s):
    print(s)
    ret = input()
    return ret
        

def user_interface():
    
    print('Интерфейс для программы тестирования')
    print('''Режимы работы: 
    1. считывание данных из файла,
    2. вывод текущих точек,
    3. вывод подробной информации о конкретной точке
    4. вывод всей информации
    5. тестирование точки''')

def parse_for_argparser():
    parser = argparse.ArgumentParser(description='script for middleware testing')
    parser.add_argument('-n', action ='store', dest='n', help='simple value')
    parser.add_argument('-o','--optional', type=int, default=2, help='provide an integer (default: 2)')
    args = parser.parse_args()
    print(args.n)
    print(args.optional) 
    return args 

def __init__():
    args = parse_for_argparser()
    s = int(user_interface() )
    token = fetchToken()
    filename = 'setting1.xlsx'
    
    settings = read_settings(filename='setting1.xlsx', sheet_name='settings')        
    parsed_data = read_data(filename='setting1.xlsx',settings = settings , sheet_name='data')    
    getters = filling_getters(settings, parsed_data, getters )             
    setters = filling_setters(settings, setters, getters)
    getters
    substitution_with_setters(settings, parsed_data, getters, setters, field_for_subst=['post_data', 'headers_value'])            
    write_2_excel_parsed_data(parsed_data, settings )


    
    
    
    token = fetchToken()
    # quirks = {'mode': 'change', 'change': 'url', 'change_type': 1, 'start_chr':'0', 'step_chr':'1', 'stop_chr':'150' }
    quirks = {'mode': 'content', 'content': 'header','change_type': 1,'step':10, 'chr':'a', 'content_num': 100}
    verbs = ['get', 'head', 'options', 'patch','connect', 'delete', 'post', 'put', 'trace']
    points = ['token', 'settings', 'links', 'relPhone', 'countMain', 'countDpi', 'bonus', 'offer']
    log = [{}, {}]
    asyncio.run( _requests(verbs, points, token,  debug=2 ))
    _write_response_to_excel('test')
