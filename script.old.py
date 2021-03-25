from json.decoder import JSONDecodeError
from typing import Dict
import aiohttp
import xlsxwriter
import json, yaml
import requests, asyncio
from openpyxl import load_workbook

           
def ifIter(v):
    dictionary = {1:1}
    lists = [0]
    if type(v) == type(dictionary) or type(v) == type(lists):
        return 1
    else:
        return 0
    
    
def ifList(v):
    lists = [0]
    if type(v) == type(lists) and len(v)>0 :
        return 1
    else:
        return 0
  
        
def ifDict(v):
    dictionary = {1:1}
    if type(v) == type(dictionary) and len(v)>0:
        return 1
    else:
        return 0
    
    
def printDict(v0):
    if ifDict(v0):
        for k1,v1 in v0.items():
            return(k1, v1)
    elif ifList(v0):
        for v1 in v0:
            return(v1)
    else:
        return(v0)
         
         
def random_chr(n,m=0,k=1, binary=1):
    if binary:
        rand=''
        if m==-1:
            n,m=m,n
        for i in range(k):
            rand+= chr(n+i)
        return rand
    return chr(n)
    
    
def url_creator(point, url_d={}, debug=0):
    if url_d!={}:
        
                    
        pass
    else:
        
        mw= "mw-tst.itsmartflex.com/" 
        #mw = "mw-apitst.vodafone.ua/" #mw_v
        
        phone_rp = "380662584192"
        phone_сs= "380662583046"
        phone_b= "380662583045"
        phone_o= "380662584418"
        product_id = "?product.id="
        siebel_token = "?siebelToken=177C50397B39937400431DC33E982F3D"
        
        point_s = "MYVF-SETTINGS"
        point_l = "MYVF-LINKS"
                
        get_token = "uaa/oauth/token?grant_type=client_credentials"
        relat_phone = "customer/api/customerManagement/v3/customer/"
        counters_i_bonus = "prepaybalance/tmf-api/prepayBalanceManagement/v4/bucket"
        offer = "qualification/api/productOfferingQualificationManagement/v1/productOfferingQualification/"
        setting_i_links = "entity/api/functions/"
        
        proto = "https"
        slash = "://"
        
    if (point=="token"):
        url = proto + slash + mw + get_token
    if (point=="relPhone"):
        url = proto + slash + mw + relat_phone + phone_rp
    if (point=="countMain" or point=="countDpi"):
        url = proto + slash + mw + counters_i_bonus + product_id + phone_сs
    if (point=="bonus"):
        url = proto + slash + mw + counters_i_bonus + '/' + phone_b
    if (point=="offer"):
        url = proto + slash + mw + offer + phone_o + siebel_token
    if (point=="settings"):
        url = proto + slash + mw + setting_i_links + point_s
    if (point=="links"):
        url = proto + slash + mw + setting_i_links + point_l
        
    log_print(debug, 2, 'url_creator ',url, point)
    
    return url


def header_creator(point, token='', debug=0, fil_type = 0, filler= ''):
 
    log_print(debug, 2, 'heador_creator', filler)
    filler0 = filler if fil_type>=0 else ''
    filler1 = filler if fil_type>=1 else ''
    filler2 = filler if fil_type>=2 else ''
    filler3 = filler if fil_type>=3 else ''
    filler4 = filler if fil_type>=4 else ''
    filler5 = filler if fil_type>=5 else ''
    filler6 = filler if fil_type>=6 else ''
    filler7 = filler if fil_type>=7 else ''
    lang = ['Accept-Language'+filler6 , 'ru'+filler0]
    cont = ['Content-Type'+filler5 , "application/json"+filler1]
    auth = ['Authorization'+filler4 , 'Basic aW50ZXJuYWw6aW50ZXJuYWw='+filler2 ]
    auth2 = ['Authorization'+filler4 , 'Bearer ' + token +filler2 ]
    prof = ['Profile' + filler7] 
    if (point=="token"):
        headers = {lang[0] : lang[1],
                   cont[0] : cont[1],
                   auth[0] : auth[1]  }
    if (point=="settings" or point=="links"):
        headers = {lang[0] : lang[1],
                   cont[0] : cont[1],
                   auth2[0] : auth2[1] }
    if (point=="relPhone"):
        headers = {prof[0] : 'RELATED-PARTY' + filler3,
                   lang[0] : lang[1],
                   cont[0] : cont[1],
                   auth2[0] : auth2[1] }
    if (point=="countMain"):
        headers = {prof[0]: 'COUNTERS' + filler3,
                   lang[0] : lang[1],
                   cont[0] : cont[1],
                   auth2[0] : auth2[1] }
    if (point=="countDpi"):
        headers = {prof[0]: 'COUNTERS-DATA' + filler3,
                   lang[0] : lang[1],
                   cont[0] : cont[1],
                   auth2[0] : auth2[1] }
    if (point=="bonus"):
        headers = {prof[0]: 'BONUS' + filler3,
                   lang[0] : lang[1],
                   cont[0] : cont[1],
                   auth2[0] : auth2[1] }
    if (point=="offer"):
        headers = {prof[0]: 'ACTIVE-OFFER' + filler3,
                   lang[0] : lang[1],
                   cont[0] : cont[1],
                   auth2[0] : auth2[1] }
    
    log_print(debug, 4,'header_creator2',point)
    log_print(debug, 5,'header_creator3',headers)
    return headers
         
              
def data_creator(s, data = {}):
    if data:    
        for d in data:
            data = {d +  s : data[d] + s} 
    else:
            data = {s : s} 
    return data
 
 
def manageRequest(point, verb , token=None, headers={}, data={}, debug=0, log={} ):
    if token==None:
        url = url_creator('token', {}, debug)
        headers = header_creator('token', '', debug)
        r = requests.request(method = 'post', url= url, headers=headers)
        # r = _requests(url, 'post', headers, data, debug)
        token = r.json()["access_token"]
        log_print(debug, 4,'manageRequest token', token)
    
    url = url_creator(point, {}, debug)
    headers = header_creator(point, token, debug)
    r = requests.request(method = verb, url= url, headers=headers)
    # r = _requests(url, verb, headers, data, debug)    
    return (r, verb, url, point, r.json())    



def many_req_gen(verbs, points, token, data={}, debug=0 ):
    
    for point in points:
        log_print(debug, 1, 'many_rew_gen0', point, 'gen' )
        for verb in verbs:
            log_print(debug, 2, 'many_rew_gen1', verb, 'gen' )
            url = url_creator(point, {}, debug)
            headers = header_creator(point, token, debug)        
            x = (verb, url, point, headers, data, debug)
            log_print(debug, 3,'many_rew_gen2', x)
            log_print(debug, 4,'many_rew_gen3', 'yield')
            yield x


def create_filler(n, s='a'):
    st = ''
    for i in range(n):
        st+= s
    return st


def log_print(debug, n, *message):
    if debug>= n:
        print(*message)

 
def gen_u(point, quirks, debug=0):
    if quirks['change'] == 'url':
        if quirks['mode'] == 'content':
            filler = quirks['chr'] * quirks['content_num']
            for i in range(quirks['step']):
                url = url_creator(point, {}, debug) + filler
                filler = filler + quirks['chr']* quirks['content_num']
                log_print(debug, 4, 'gen_u', url)
                yield url
                
        elif quirks['mode'] == 'change':
            for i in quirks['stop'] - quirks['start']:
                filler =  chr( [quirks['start'] +i ] ) 
                url = url_creator(point, {}, debug) + filler
                log_print(debug, 4, 'gen_u', url)
                yield url
    else:
        while True:
            yield url_creator(point, {}, debug)
 
 
 
def gen_h(point, token, quirks, debug=0):
    log_print(debug, 4, 'gen_h 0', quirks)
    if quirks['change'] == 'header':
        log_print(debug, 4, 'gen_h 1')
        if quirks['mode'] == 'content':
            log_print(debug, 4, 'gen_h 2')
            filler = quirks['chr'] * quirks['content_num']
            for i in range(quirks['step']):
                log_print(debug, 4, 'gen_h 3')
                header = header_creator(point, token, debug, quirks['change_type'], filler)
                log_print(debug, 4, 'gen_h3.5 lenof filler', len(filler))
                filler = filler + quirks['chr']* quirks['content_num']
                log_print(debug, 4, 'gen_h 4',header)
                yield header
                
        elif quirks['mode'] == 'change':
            for i in quirks['stop'] - quirks['start']:
                filler =  chr( [quirks['start'] +i ] ) 
                header = header_creator(point, token, debug, quirks['change_type'], filler)
                log_print(debug, 4, 'gen_h6', header)
                yield header
    else:
        while True:
            yield header_creator(point, token, debug)


def gen_d(quirks, data={}, debug=0):
    if quirks['change'] == 'data':
        log_print(debug, 4, 'gen_d 0')
        if quirks['mode'] == 'content':
            log_print(debug, 4, 'gen_d 1')
            filler = quirks['chr'] * quirks['content_num']
            for i in range(quirks['step']):
                log_print(debug, 4, 'gen_d 2')
                data = data_creator(filler, data) 
                filler = filler + quirks['chr']* quirks['content_num']
                log_print(debug, 4, 'gen_d 4', data)
                yield data
                
        elif quirks['mode'] == 'change':
            for i in quirks['stop'] - quirks['start']:
                data = data_creator(filler, data) 
                filler =  chr( [quirks['start'] +i ] ) 
                log_print(debug, 4, 'gen_d', data)
                yield data
    else:
        while True:
            yield data
     
           
def many_req_quirks_gen(verbs, points, token, quirks, data, debug ):
    
    log_print(debug, 4,' many_req_quirks_gen quirks mode')
    for point in points:
        for verb in verbs:
            subgen_h = gen_h(point, token, quirks, debug)
            subgen_u = gen_u(point, quirks, debug)
            subgen_d = gen_d(quirks, data, debug)
            if quirks['mode'] == 'change':
                n = (quirks['stop_chr'] - quirks['start_chr'] )/ quirks['step_chr']
            if quirks['mode'] == 'content':
                n = quirks['step']
                 
            for i in range(n):
                header = next(subgen_h)
                url = next(subgen_u)
                data = next(subgen_d)
                x = (verb, url, point, header, data, i)

                log_print(debug, 4, ' many_req_quirks_gen', i,n)
            yield x                       
   
       
async def _requests(verbs, points, token, quirks=None, headers = {}, data= {}, debug=0):
    if quirks is not None:
        gen = many_req_quirks_gen(verbs, points, token, quirks, data, debug )
        log_print(debug,2,'_requests1')
    else:
        gen = many_req_gen(verbs, points, token, data, debug )
        log_print(debug,2,'_requests2')
    log_print(debug,2,'_requests3')
    async with aiohttp.ClientSession() as session:
        log_print(debug,2,'_requests session')
        q=0
        try:
            while x:= next(gen):
                q+=1
                verb, url, point, headers, data, filler = x #next(gen)
                async with session.request(verb, url, headers=headers, data=data) as response:
                # async with session.post(url, headers=headers) as response:
                    status =  response.status
                    log_print(debug, 1, '_requests, in session', status)
                    try:
                        resp = await response.json()
                    except aiohttp.client_exceptions.ContentTypeError:
                        log_print(debug, 1, '_requests not a json, q', q)
                        resp = await response.text()

                    if status>=200 and status<300:
                        log[0][(point, url, verb, filler)] = (status, headers, data, resp)
                    else:
                        log[1][(point, url, verb, filler)] = (status, headers, data, resp)
            await session.close()
        except StopIteration as e:
            log_print(debug, 1,'_requests', e, q,filler)


def fetchToken(url=None, headers=None, mw = None):
    if mw is None:
        mw = 'mw-tst.itsmartflex.com'
    elif mw == 'vodafone':
        mw = 'mw-apitst.vodafone.ua' #mw-tst.itsmartflex.com
    
    if url is None:
        url = 'https://' + mw + '/uaa/oauth/token?grant_type=client_credentials'
        url2 = 'https://' + mw + '/uaa/oauth/token?grant_type=password'
    if headers is None:
        basic1 = 'aW50ZXJuYWw6aW50ZXJuYWw='
        basic2= 'c3ZjLXNpZWJlbC1tdzoxMjM0='
        headers = {"Authorization": "Basic " + basic1,"Content-Type": "application/json"}
        headers2 = {"Authorization": "Basic c3ZjLXNpZWJlbC1tdzoxMjM0=","Content-Type": "application/x-www-form-urlencode"}
    # data = {'username' :'380952240016'}
    print(url)
    print(headers)
    r = requests.post(url, headers = headers)#, data = data) 
    print(r.status_code)
    json_response = r.json()
    return json_response['access_token']


def fetchPass(token = None, url=None, headers=None, mw = None):
    if mw is None:
        # mw = 'mw-apitst.vodafone.ua'
        mw = 'mw-tst.itsmartflex.com'
    if url is None:
        # url = 'https://' + mw + '/uaa/oauth/token?grant_type=client_credentials'
        url = 'https://' + mw + '/uaa/oauth/token?grant_type=password'
    if token is None:
        token = 'eyJhbGciOiJSUzI1NiIsInR5cCI6IkpXVCJ9.eyJzY29wZSI6WyJtc191YWFfY3JlZGVudGlhbF90eXBlX3Bhc3N3b3JkX2p3dCIsIm9wZW5pZCJdLCJleHAiOjE2MTE3NjI5NzksImFkZGl0aW9uYWxEZXRhaWxzIjp7Im1zaXNkbiI6IjM4MDk1MjI0MDAxNiJ9LCJhdXRob3JpdGllcyI6WyJST0xFX1NZU1RFTSJdLCJqdGkiOiI3NmRiNTEyMS0zZjJmLTQzZTMtYTJjZi1lMTA4NGNjM2YzZTciLCJ0ZW5hbnQiOiJYTSIsImNsaWVudF9pZCI6InN2Yy1zaWViZWwtbXcifQ.d50VVTF50QF-nqvxrmzcXWg9Ih0qRdTlPYdfhi_RKzgDL_RjDtu_sIiv6FWLKGyAgYcpVD-ukL2F8BU53OGAqgEbJo0zJLzVXEIUDcdc-zxZufPf-x6MtFhwSo4Xk6v22esSgPtU-LD2Kk8wfr4_yA0BChu4OXS2uXwvy7qaQS7h889kIyO2pSeHBJdtS0-EjrP4cFebd32hFZ9J_25ockKjJ17y08ubYQh_YMrbOzHlWrfmbYhMVN71uadgOGfRnwkSWzVJ-GThcFasXTuK7ONKa748sCQWG_q4R7QEzZbvDnFk2Dw-VXyb73kOJQKOJY3QblDTVmkqF-yg8DMKGQ'

    if headers is None:
        # headers = {"Authorization": "Basic aW50ZXJuYWw6aW50ZXJuYWw=","Content-Type": "application/json"}
        # –header ‘Content-Type: application/x-www-form-urlencoded’
        headers = {"Authorization": "Basic YXBwLW15dm9kYWZvbmUtbXc6UGg1ZDg2QnpCNg==","Content-Type": "application/x-www-form-urlencoded", 'Profile': 'MYVODAFONE'}
        headers2 = {"Authorization": "Basic c3ZjLXNpZWJlbC1tdzoxMjM0’=","Content-Type": "application/x-www-form-urlencode"}
    data = {'username' :'380952240016', 'password' : token}

    print(url)
    print(headers)
    r = requests.post(url, headers = headers, data = data) 
    print(r.status_code)
    try:
        json_response = r.json()['access_token']
    except:
        json_response = r.text
        print(json_response)
    return json_response


def print_elem(dic, key):
    if type(dic) == type(dictionary) and dic.get(key, 0):
        return key
    else:
        return key + ' ' + dic[key]


def isType(e):  
    '''-1 None, 0 if empty array, 1 if str int float, 2 if list, 3 if dict
    '''
    if e is None:
        return -1
    if e == [] or e == {} or e == ():
        return 0
    if str(type(e)) == "<class 'bool'>" or str(type(e)) == "<class 'int'>" or str(type(e)) == "<class 'float'>" or str(type(e)) == "<class 'str'>":
        return 1 
    if str(type(e)) == "<class 'list'>":
        return 2 
    if str(type(e)) == "<class 'dict'>":
        return 3  


def is_int(i):
    try:
        int(i)
        return 1 
    except:
        return 0
    
    
def none_safe_str(s):
    if s is None:
        return s
    else:
        return str(s)
    
    
def checkArray(el): #0 str, 1 list, 2 dict, 3 listdict, 4 listlist, 5 dictlist, 6 dictdict
    if isType(el)<2: #isType(): -1 None, 0 if empty array, 1 if str int float, 2 if list, 3 if dict
        return 0
    for e in el:
        # print('checkarray0',isType(e),type(e))
        # print(e)
        if isType(el)==2: #list
            if isType(e)>=2: #list or dict
                return isType(e) + 1
            else: return 1
        if isType(el)==3: #dict
            if isType(el[e])>=2: #list or dict
                return isType(el[e]) + 3
            else: return 2
    return 0        
        
        
def parse_string(s):
    for c in s:
        if c=='[':
            pass
        if c=='{':
            print()
        # if 


def parse_str(s,n, dic=0):
    if isType(s)<2:
        yield (s,n,dic)
    if isType(s)==2:
        for c in s:
            yield from parse_str(c, n+1, dic)
        print()
    if isType(s)==3:
        for c in s:
            yield (c,n+1,dic)
            yield from parse_str(s[c], n, dic+1)
        dic -=1
    

def worksheet_write_twice_shift(sheet, x,y, s, shift, s_new, param='', x0=0, y0=0, type_s=0): #i,d
    if param=='':
        sheet.write(x,y, str(s))
        print(x,y,shift)
        if s_new!='':
            print('worksheet_write_twice_shift s_new no param')
            print('||||||||||||||||||||||||||||||||||')
            print(s_new)
            print('||||||||||||||||||||||||||||||||||')
            sheet.write(x,y+shift, str(s_new))
    else:
        sheet.write(x,y, str(s), param)
        print(x,y,shift)
        if s_new!='':
            print('worksheet_write_twice_shift s_new param')
            print('||||||||||||||||||||||||||||||||||')
            print(s_new)
            print('||||||||||||||||||||||||||||||||||')
            sheet.write(x,y+shift, str(s_new), param)
    if x0!=0 or y0!=0 and s_new!='':
        if s==s_new: check = 1
        else: check = 0
        
        if param=='':
            if type_s: sheet.write(x0,y0+1, type(check) )
            sheet.write(x0,y0, str(bool(check)) )
        else:
            if type_s: sheet.write(x0,y0+1, type(check) )
            sheet.write(x0,y0, str(bool(check)), param)
    else:
        check = 0    
    return check


def _write_response_to_excel(resp, resp_new='',shift=4, filename='test.xlsx', verbose=10, workbook='', current_worksheet='',new_worksheet='testings',border_draw=1,x0=0,y0=0):
    '''Функция для красивого вывода ответа на запрос в файл екселя
    '''
    gen = parse_str(resp,0)
    if resp_new !='':
        print('++++++++++++++++++++++')
        gen_example = parse_str(resp_new,0)
        
    if workbook=='':
        workbook = xlsxwriter.Workbook(filename)
        need_close = 1
    else:
        need_close =0
        
        
    if current_worksheet=='':
        worksheet = workbook.add_worksheet(new_worksheet)
    else:
        worksheet = current_worksheet
    if border_draw ==1:
        border = workbook.add_format()
        border.set_border(style=1)

        border_no_top = workbook.add_format()
        border_no_top.set_bottom()
        border_no_top.set_left()
        border_no_top.set_right()
        border_no_topright = workbook.add_format()
        border_no_topright.set_border(style =3)
        border_no_topright.set_bottom()
        border_no_topright.set_left()
        border_no_left = workbook.add_format()
        border_no_left.set_border(style =3)
        border_no_left.set_bottom()
        border_no_left.set_top()
        border_no_left.set_right()
        border_no_right = workbook.add_format()
        border_no_right.set_border(style =3)
        border_no_right.set_bottom()
        border_no_right.set_top()
        border_no_right.set_left()
        border_no_bottom = workbook.add_format()
        border_no_bottom.set_top()
        border_no_bottom.set_left()
        border_no_bottom.set_right()
        border_no_bottomleft = workbook.add_format()
        border_no_bottomleft.set_border(style =3)
        border_no_bottomleft.set_top()
        border_no_bottomleft.set_right()

    try:
        i=0
        shift_right_prev=0
        shift_right_max = 0
        str_resp_old=''
        q=0
        border_prev = ''
        while True:
            str_resp,shift_down,shift_right =next(gen)
            if resp_new!='':
                str_resp_old,shift_down_old,shift_right_old =next(gen_example)  #str_resp_prev,n2,d2 =next(gen_example) 
                # print('*********************')
                # print(s2)
                # print('*********************')
                
            else: str_resp_old=''
            shift_right+=y0
            if i==0:
                shift_right_prev=shift_right
                n_prev=shift_down
                     
            if q==0:
                i =x0
                shift_right_prev=shift_right
                i_prev=i
            if shift_right>shift_right_prev:
                i-=1
            
            if shift_down == n_prev and q>0: #если уровень списка одинаков для текущего и предыдущего
                if shift_right==shift_right_prev: #если уровень словаря одинаков для текущего и предыдущего
                    if border_prev == 'no left': #если для предыдущей клеточки установлена граница "без левой стороны", то устанавливаем "без нижней и левой"
                        worksheet.write(i_prev, shift_right_prev, str_resp(str_resp_old),border_no_bottomleft)
                            
                    else: #иначе устанавливаем только  "без нижней"
                        worksheet.write(i_prev, shift_right_prev, str_resp(str_resp_old),border_no_bottom)

                    # sett.write(i, d, str(s),border_no_top)
                    worksheet_write_twice_shift(worksheet, i, shift_right,str_resp, shift, str_resp_old, param = border_no_top, x0=i, y0=y0-1) #i,d
                    border_prev = 'no top' #устанавливаем "без верхней " для текущей
                    q+=1
                elif shift_right>shift_right_prev:
                    if border_prev == 'no top': #если для предыдущей клеточки установлена граница "без правой стороны", то устанавливаем "без верхней и правой"
                        worksheet.write(i_prev, shift_right_prev, str_resp(str_resp_old),border_no_topright)
                    else: #иначе устанавливаем только "без правой"
                        worksheet.write(i_prev, shift_right_prev, str_resp(str_resp_old),border_no_right)
                    # sett.write(i, d, str(s),border_no_left)
                    worksheet_write_twice_shift(worksheet, i, shift_right,str_resp, shift, str_resp_old, param = border_no_left, x0=i, y0=y0-1) #i,d
                    border_prev = 'no left'
                    q+=1
                else:
                    # sett.write(i, d, str(s))
                    worksheet_write_twice_shift(worksheet, i, shift_right,str_resp, shift, str_resp_old, x0=i, y0=y0-1) #i,d
                    border_prev = ''
                    q+=1
            else:
                # sett.write(i, d, str(s))
                worksheet_write_twice_shift(worksheet, i, shift_right,str_resp, shift, str_resp_old, x0=i, y0=y0-1) #i,d
                border_prev = ''
                q+=1
            
            shift_right_prev = shift_right
            shift_right_max = max(shift_right_max, shift_right)
            i_prev = i
            n_prev = shift_down
            str_resp_old = str_resp
            i+=1
    except Exception as e:
        print(e)
    finally:
        if need_close:
            workbook.close()
    return (workbook, worksheet, i-x0, shift_right_max - y0)


def check_if_cell_is_not_empty(s):
    try:
        if s is not None and s != '':
            return 1
        else:
            return 0
    except Exception as e:
        print(e)
       

def append_list_in_dict(dict_, list_, elem):
    print(dict_, list_, elem)
    try:
        dict_[list_].append(elem)
    except:
        dict_[list_] = []
        dict_[list_].append(elem)
    return  dict_

def add_record_in_dict(dict_, key, key2, elem):
    print(dict_, key, key2, elem)
    try:
        dict_[key][key2] = elem
    except:
        dict_[key] = {}
        dict_[key][key2] = elem
    return  dict_
      
      
parsed_settings['setter'][c[0].value] = {c[2].value:c[1].value}
append_list_in_dict(parsed_settings['field of response'], c[1].value, c[2].value )
filename = 'setting1.xlsx'

      
settings = read_settings(filename='setting1.xlsx', sheet_name='settings')        
  
settings      
       
def parse_settings_in_read_settings(settings, c, mode):
    if c[0].value == 'начальная точка': settings['diapasone'].append(c[1].value)
    if c[0].value == 'конечная точка': settings['diapasone'].append(c[1].value)
    if c[0].value == 'уровень раскрытия ответа запроса': settings['verbose_level'] = int(c[1].value)   
    if c[0].value == 'url для токена миддлваре': url_for_token_mv = c[1].value
    if c[0].value == 'генерация токена миддлваре': gen_for_token_mv = c[1].value
    if c[0].value == 'url для токена сибеля':  url_for_token_siebel = c[1].value            
    if c[0].value == 'генерация токена сибеля':  gen_for_token_siebel = c[1].value            
    if c[0].value == 'проверка определения способа подстановки токена': url_for_token_check = c[1].value
    if c[0].value == 'токен миддлваре': 
        if c[1].value == 'из точки':
            settings['url_mw']['from'] = 'from_point_and_url'            
            settings['url_mw']['from_point'] = c[2].value            
            settings['url_mw']['from_url'] = url_for_token_mv           
            settings['url_mw']['how_gen'] = gen_for_token_mv     
            settings['url_mw']['how_check_url'] = url_for_token_check     
        if c[1].value == 'взять из клетки':
            settings['url_mw']['from'] = 'from_cell'            
            settings['url_mw']['token'] = c[2].value            
        if c[1].value == 'ничего не делать':
            settings['url_mw']['from'] = 'nothing'            
    if c[0].value == 'токен сибель': 
        if c[1].value == 'из точки':
            settings['url_siebel']['from'] = 'from_point_and_url'            
            settings['url_siebel']['from_point'] = c[2].value            
            settings['url_siebel']['from_url'] = url_for_token_siebel           
            settings['url_siebel']['how_gen'] = gen_for_token_siebel           
            settings['url_siebel']['how_check_url'] = url_for_token_check     
        if c[1].value == 'взять из клетки':
            settings['url_siebel']['from'] = 'from_cell'            
            settings['url_siebel']['token'] = c[2].value            
        if c[1].value == 'ничего не делать':
            settings['url_siebel']['from'] = 'nothing'            

    if c[0].value == 'старый токен': settings['verbose_level'].append(c[1].value)            
    if c[0].value == 'проверка совпадения ответа': 
        if c[1].value == 'простая': 
            settings['how to check response'] ='simple'            
        if c[1].value == 'сложная': 
            settings['how to check response'] = 'complicated'            
            if c[2].value == 'исключить': 
                settings['exclude or include fields'] = 'exclude'            
                mode = 'mode_exc_or_inc_values_in_cheking_response'        
            if c[2].value == 'включить': 
                settings['exclude or include fields'] = 'include'    
                mode = 'mode_exc_or_inc_values_in_cheking_response'        
    if c[0].value == 'вывод типов полей в ответе': settings['show type of field in response'] = c[1].value            
    if c[0].value == 'за каждым элементом списка ответов перевод строки ': settings['enter after each element of list'] = c[1].value            
    if c[0].value == 'ответ в рамочки': settings['response in borders'] = c[1].value            
    if c[0].value == 'размер столбцов': 
        mode = 'mode_parse_columns_width'
        settings['columns'][c[1].value] = c[2].value            
    if c[0].value == 'получение значений' : 
        if c[1].value == 'да':
            mode= 'mode_parse_for_getter'
            
    if c[0].value == 'подстановка значений' : 
        if c[1].value == 'да':
            mode= 'mode_parse_for_setter'
            
    return settings, mode
        
       
def read_settings(filename, sheet_name='settings', diapasone=('A1', 'C100'), validate_function=None, **kwargs ):
    
    wb = load_workbook(filename = filename)
    if sheet_name == '__active':
        sheet = wb.active
    else: 
        sheet = wb.get_sheet_by_name(sheet_name)
    try:
        cells = sheet[diapasone[0]: diapasone[1]]
    except Exception as e:
        printlog('Неверный диапазон значений для парсинга')
    c = [0 for i in range(len(cells) )]
    parsed_settings = {'diapasone' : [], 'verbose_level' : 0, 'url_mw' : {}, 'url_siebel' : {}, 'how to check response' : '', 'exclude or include fields': '',
                            'field of response': {},'show type of field in response' : '', 'enter after each element of list': '', 'response in borders':'', 
                            'columns' : {}, 'token-mw' :'', 'token-siebel' :'', 'getter':{}, 'setter':{}, 'getter_inv':{}, 'setter_inv':{}}

    mode = ''
    i = 0
    for *c, in cells:  
        print(c[0].value,c[1].value,c[2].value)
                
        if check_if_cell_is_not_empty(c[0].value):

            if not (mode == 'mode_parse_for_setter' or mode == 'mode_parse_for_getter'):
                parsed_setings, mode = parse_settings_in_read_settings(parsed_settings, c, mode)
            else:
                if mode == 'mode_parse_for_getter':
                    # parsed_settings['getter'][c[0].value] = {c[2].value:c[1].value}
                    add_record_in_dict(parsed_settings['getter'], c[0].value, c[2].value, c[1].value)
                    add_record_in_dict(parsed_settings['getter_inv'], c[0].value, c[1].value, c[2].value)
                if mode == 'mode_parse_for_setter':
                    # parsed_settings['setter'][c[0].value] = {c[2].value:c[1].value}
                    add_record_in_dict(parsed_settings['setter'], c[0].value, c[2].value, c[1].value)
                    add_record_in_dict(parsed_settings['setter_inv'], c[0].value, c[1].value, c[2].value)


        else:
            if mode == 'mode_parse_for_setter' or mode == 'mode_parse_for_getter':
                mode = ''
            
            if check_if_cell_is_not_empty(c[1].value):
            
                if mode == 'mode_exc_or_inc_values_in_cheking_response':
                    append_list_in_dict(parsed_settings['field of response'], c[1].value, c[2].value )

                if mode == 'mode_parse_columns_width':
                    parsed_settings['columns'][c[1].value] = c[2].value            
        
    return parsed_settings
        

def parse_2_dict(*args):
    s = ''
    i=1
    if len(args)==1:
        return args[0]
    for a in args:
        if i==1:
            i+=1
            continue
        s += '{"'+a+'":'
        i+=1
    for j in range(i):
        s+='}'
    return s


token = fetchToken()
resp_sett = requests.get(url_creator('token'), header_creator('token'))
resp_sett
gettoken
token
 
 
def json_parse(s):
    if s=='':
        return s
    try:
        str_json = json.loads( s.replace("'",'"'))
    except:
        try:
            q0 = s.find('"')
            q1 = s.rfind('"')
            print(q0,q1)
            print(s)
            s = s[:q0]+ s[q0:q1].replace("'",  '') +s[q1:]
            s = s.replace("'", '"')
            print(s)
            str_json = yaml.load(s)
        except Exception as e:
            print(e)
            str_json = 'json decode error, maybe unexpectend of string'
    return str_json
 
        
def read_data(filename, sheet_name='__active', diapasone=('A1', 'I40'), validate_function=None, **kwargs ):
    
    wb = load_workbook(filename = filename)
    if sheet_name == '__active':
        sheet = wb.active
    else: 
        sheet = wb[sheet_name]
    try:
        cells = sheet[diapasone[0]: diapasone[1]]
    except Exception as e:
        print('Неверный диапазон значений для парсинга')
    c = [0 for i in range(len(cells) )]
    parsed_data = {'points' : [], 'verbs' : {}, 'headers_name' : {}, 'headers_value' : {}, 'parsed_headers_value' : {}, 'url' : {}, 
                   'post_data':{}, 'parsed_post_data':{}, 'response' : {}, 'parsed_response' : {}, 'response_code' : {}, 'token' :0}

    i = 0
    for *c, in cells:  
        i +=1
        if i==1:
            continue
        if c[1].value is not None and is_int(c[0].value) :
            parsed_data['points'].append(c[1].value)
            last_point = c[1].value
            parsed_data['verbs'][c[1].value] = []
            parsed_data['verbs'][c[1].value].append(c[2].value)
            parsed_data['headers_name'][c[1].value] = []
            parsed_data['headers_name'][c[1].value].append(c[3].value)
            parsed_data['headers_value'][c[1].value] = {}
            parsed_data['headers_value'][c[1].value][c[3].value] = c[4].value
            parsed_data['parsed_headers_value'][c[1].value] = {}
            s = c[4].value
            try:
                str_json = json_parse(s)
            except Exception as e:
                print(e)
                str_json = ''
            parsed_data['parsed_headers_value'][c[1].value][c[3].value] = str_json
            parsed_data['url'][c[1].value] = []
            parsed_data['url'][c[1].value].append(c[5].value)
            parsed_data['post_data'][c[1].value] = {}
            parsed_data['post_data'][c[1].value][c[2].value] = c[6].value
            parsed_data['parsed_post_data'][c[1].value] = {}
            s = c[6].value
            try:
                str_json = json_parse(s)
            except Exception as e:
                print(e)
                str_json = ''
            parsed_data['parsed_post_data'][c[1].value][c[2].value] = str_json                
            
            parsed_data['response_code'][c[1].value] = {}
            parsed_data['response_code'][c[1].value][c[2].value] = c[7].value
            parsed_data['response'][c[1].value] = {}
            parsed_data['response'][c[1].value][c[2].value] = c[8].value                
            parsed_data['parsed_response'][c[1].value] = {}
            s = c[8].value
            try:
                str_json = json_parse(s)
            except Exception as e:
                print(e)
                str_json = ''
            parsed_data['parsed_response'][c[1].value][c[2].value] = str_json                
                
        elif last_point is not None:
            if c[2].value is not None:
                parsed_data['verbs'][last_point].append(c[2].value)
            if c[3].value is not None:
                parsed_data['headers_name'][last_point].append(c[3].value)
                if c[4].value is not None:
                    parsed_data['headers_value'][last_point][c[3].value] = c[4].value
            if c[5].value is not None:
                parsed_data['url'][last_point].append(c[5].value)
            if c[6].value is not None:
                parsed_data['post_data'][last_point].append(c[6].value)
            if c[7].value is not None:
                parsed_data['response_code'][last_point][c[2].value] = c[7].value
            if c[8].value is not None:
                parsed_data['response'][last_point][c[2].value] = c[8].value
                s = c[8].value
                try:
                    str_json = json_parse(s)
                except Exception as e:
                    print(e)
                    str_json = ''
                parsed_data['parsed_response'][last_point][c[2].value] = str_json  
            
    return parsed_data
        


parsed_data = read_data(filename, sheet_name='data')    

for p in parsed_data['headers_name']:
    print(parsed_data['headers_name'][p])
    


for p in parsed_data['headers_value']:
    print(parsed_data['headers_value'][p])


print(parsed_data['parsed_headers_value'])

parsed_data['headers_value']

s='{"error": "method_not_allowed", "error_description": "Request method "GET" not supported"}'
json.loads(s)


settings = read_settings(filename='setting1.xlsx', sheet_name='settings')        
settings


parsed_data['headers_value']
parsed_data['points']
parsed_data['post_data']
list(settings['getter'].keys())
list(parsed_data['response']['token'].items())
list(parsed_data['parsed_response']['token'].items())
list(settings['getter'].items())

list(parsed_data['parsed_response']['token'].items())
parsed_data['parsed_response']['token']

for i in parsed_data['parsed_response']['token']:
    print(i)
    print(parsed_data['parsed_response']['token'][i])
    
list(parsed_data['response']['token'].items())

setters = {}
getters = {}

for points in parsed_data['parsed_response']:
    for verbs in parsed_data['parsed_response'][points]:
        print(points, verbs)
        print(parsed_data['parsed_response'][points][verbs])
        print(type(parsed_data['parsed_response'][points][verbs]))
        print()
        print()
        
        for keys in parsed_data['parsed_response'][points][verbs]:


            if keys in settings['getter']:
                for m in settings['getter'][keys]:
                    print('point=',points, verbs, keys, '  var=', m, '  ref=', settings['getter'][points][m])
                    if settings['getter'][verbs][m] in parsed_data['parsed_response'][points]:
                        print('|||||||||')
                    
            # print(hv, m)
            

'token_mw' in dict(settings['getter']['token'].items())
getters = filling_getters(settings, parsed_data, getters )             
setters = filling_setters(settings, setters, getters)
setters
settings['getter']
parsed_data['parsed_response']
                
def filling_getters(settings, parsed_data, getters ):             
    
    for points in parsed_data['parsed_response']:
        if points in settings['getter'].keys():
            for verbs in parsed_data['parsed_response'][points]:
                if isType(parsed_data['parsed_response'][points][verbs]) >=2:
                    for r in parsed_data['parsed_response'][points][verbs]:
                        if isType(parsed_data['parsed_response'][points][verbs][r]) >=2:
                            for re in parsed_data['parsed_response'][points][verbs][r]:
                                if isType(parsed_data['parsed_response'][points][verbs][r][re]) >=2:
                                    for res in parsed_data['parsed_response'][points][verbs][r][re]:
                                        if settings['getter'].get(points,0):
                                            if res in list(settings['getter_inv'][points].keys()):
                                                alias = list(settings['getter'][points].keys())[0]
                                                getters[alias] = parsed_data['parsed_response'][points][verbs][r][re][res]
                                                
                                else:
                                    if settings['getter'].get(points,0):
                                        if re in list(settings['getter_inv'][points].keys()):
                                            alias = list(settings['getter'][points].keys())[0]
                                            getters[alias] = parsed_data['parsed_response'][points][verbs][r][re]
 
                        else:
                            if settings['getter'].get(points,0):
                                if r in list(settings['getter_inv'][points].keys()):
                                    alias = list(settings['getter'][points].keys())[0]
                                    getters[alias] = parsed_data['parsed_response'][points][verbs][r]
    return getters
 
 
def filling_setters(settings, setters, getters):
    
    for sett in settings['setter']:
        for set in settings['setter'][sett]: 
            print('set= ', set)
            var = settings['setter'][sett][set]
            print('var= ', var)
            i0 = set.find('{{')
            i1 = set.find('}}')
            print('io i1 ', i0, i1)
            if i0>=0 and i1>=0 :
                print('set[i0+2:i1]', set[i0+2:i1])
                print(set[:i0], getters[set[i0+2:i1]], set[i1+2:], sep='')
                setters[set] = set[:i0] + getters[set[i0+2:i1]] + set[i1+2:]
            else:
                setters[set] = set
    return setters
 

settings['getter'].keys()
settings['getter']
settings['getter_inv']
settings['setter']
settings['setter_inv']
parsed_data['post_data']
getters
setters
settings['getter']['token']   
list(settings['setter_inv']['getRelPhone'].keys())
field_for_subst = ['post_data']
field ='post_data'
for p in parsed_data[field]:
    print(p)
    for v in parsed_data[field][p]:
        print(p)


'settings' in settings['setter'].keys()


def substitution_with_setters(settings, parsed_data, getters, setters, field_for_subst=['post_data', 'headers_value']):             


    for field in field_for_subst:
        print(1)
        for points in parsed_data[field]:
            print(2)
            if points in settings['setter'].keys():
                print(3)
                # for verbs in parsed_data[field][points]:
                    # print(14)
                if isType(parsed_data[field][points]) >=2:
                    print(4)
                    for r in parsed_data[field][points]:
                        print(5)
                        print(parsed_data[field][points])
                        print(parsed_data[field][points][r])
                        print(r)
                        print(isType(parsed_data[field][points][r]))
                        print(type(parsed_data[field][points][r]))
                        # if isType(parsed_data[field][points]) >=2:

                        if isType(parsed_data[field][points][r]) >=2 :
                            print(r, type(parsed_data[field][points]), len(parsed_data[field][points]))
                            print(6)
                            for re in parsed_data[field][points][r]:
                                print(17)
                                if isType(parsed_data[field][points][r][re]) >=2:
                                    print(18)
                                    for res in parsed_data[field][points][r][re]:
                                        
                                        print('settings[setter]', settings['setter'].keys(), points)
                                        if settings['setter'].get(points,0):
                                            if res in list(settings['setter_inv'][points].keys()):
                                                alias = settings['setter_inv'][points]
                                                # getters[alias] = 
                                                print('res=', res)
                                                print(parsed_data[field][points][r][re][res])
                                                print('alias', alias)
                                                # parsed_data[field][points][r][re][res] = alias #settings['setter_inv'][points][res]
                                                # setters[alias]
                                else:
                                    print(21)

                                    if settings['setter'].get(points,0):
                                        if re in list(*settings['setter'][points].items()):
                                            alias = list(settings['setter'][points].keys())[0]
                                            # getters[alias] = parsed_data[field][points][r][re]
                                            print('alias', alias)
                        else:
                            print(101)
                            if settings['setter'].get(points,0):
                                print(111)
                                if r in list(settings['setter'][points].items()):
                                    print(121)
                                    alias = list(settings['setter'][points].keys())[0]
                                    # setters[alias] = parsed_data[field][points][r]
                                    print('alias', alias)


    
    








if settings['getter'].get(point,0):
                    
                    
                    
                    
                    
                    
print(settings['getter'][''])
                    
settings['getter']
          
          
            
for hv in parsed_data['post_data']:
    if hv in list(settings['getter'].keys()):
        for m in list(settings['getter'][hv].keys()):
            print(hv, m)
        # print(hv)
        
        
for pd in parsed_data['post_data']:
    if pd in list(settings['setter'].keys()):
        print(pd)
    

    
print( list(settings['setter'].keys()))


for pd in parsed_data:
    print(pd)






type(parsed_data['post_data']['getToken']['post'])
res = parsed_data['response']['getlinks']['post']
res = str(res)
parsed_data['post_data']['getToken'].get('post','')
parsed_data['url']
write_2_excel_parsed_data(parsed_data, settings )

resp
type(res)
dict(res)
import json,yaml
resp =  res.replace("'", "\"")
d= yaml.load(s)
repr(res)

settings

res2 =res
res3 = json.loads(s3.replace("'", '"'))
res3.keys()
s2 = '{"loginV2": {"error": 0, "values": {"tempToken": "8DBA35BBA4C486592B1BCC0A984DC5BE"}},"identification": {"error": 0, "values": { "id":""}}}'
g = parse_str(res3, 0)
s,n,d = next(g)
s
s2
s3 = "{'access_token': 'eyJhbGciOiJSUzI1NiIsInR5cCI6IkpXVCJ9.eyJzY29wZSI6WyJvcGVuaWQiXSwiZXhwIjoxNjE2NDUwMzI4LCJhdXRob3JpdGllcyI6WyJTVVBFUi1BRE1JTiJdLCJqdGkiOiIyZmNhY2MzYy1mM2I0LTQzNmQtYTYyZC01ZmFlZjdmMGY0YjUiLCJ0ZW5hbnQiOiJYTSIsImNsaWVudF9pZCI6ImludGVybmFsIn0.AaK7IxOhYdjTTvDtx2qa1ndyhjhZN11teSECFcobBnQ8fdIBvsOZzOwqbYv5Kw7cPhh8kGxi5OKjFQqYQjiXIeZFpDf6BzKOvwBu2QUskdFxITpnuLxXrwhemXNHbsuLn1M46LOEmRqYJ8nl6QHEZyEB8SePSNZIWBHCUbGJKWIMoMxdB6PjQjnSSw0WbHsYj4petYEVn56pj3sppyxpr3vppjF3HWM7XtoNvzee9P6fLaKtwuEgDfOXbUYGRNiYPttkV5sR7UjWaEOd4zIklsu2vOetrojiteeVtUALCaUqNaYUoczfRMbvHdeucetrD5Lsi8wmN2oTbvM1Falmzg', 'token_type': 'bearer', 'expires_in': 86399, 'scope': 'openid', 'tenant': 'XM', 'jti': '2fcacc3c-f3b4-436d-a62d-5faef7f0f4b5'}"






def serching_and_substitution_variables(s, point, settings):
    if isType(settings['searching'])==3:
        for p in settings['searching']:
            if point == p:
                for value in settings['searching'][p]:
                    if value == s:
                        settings['searching'][p][s]
    
    if isType(settings['substitution'])==3:
        if s == ''





for i,di in d.items():
    print(i,di)

for di in d:
    print(type(di))
    print(di)
    print(type(d[di]))
    print(d[di])
    



def write_2_excel_parsed_data(parsed_data, settings, filename='response.xlsx', sheet_name='response', **kwargs ):
    workbook = xlsxwriter.Workbook(filename = filename)
    worksheet = workbook.add_worksheet(sheet_name)
    if settings is not None and settings.get('columns_width',0):
        for c, w in settings['columns_width'].items():
            worksheet.set_column(int(c), int(c), int(w))
    counter = 1
    point_counter = 1  
    for point in parsed_data['points']:
        worksheet.write( counter, point_counter, none_safe_str(point))
        verb_counter = counter
        print('log05')
        for verb in parsed_data['verbs'][point]:
            # worksheet.write( counter, point_counter+5, none_safe_str( parsed_data['post_data'][point].get(verb,'')) )
            y_data_counter =0
            try:
                data = json.loads(parsed_data['post_data'][point].get(verb,''))
            except:
                data = parsed_data['post_data'][point].get(verb,'')
            workbook,worksheet,resp_counter, y_data_counter = _write_response_to_excel( data, verbose=3, current_worksheet=worksheet, workbook=workbook,x0=counter,y0=point_counter+5)
            print('log0')
            worksheet.write( verb_counter, point_counter+1, str(verb))
            try:
                if verb=='post' and parsed_data['post_data'][point].get(verb,''):
                    r = requests.request(verb, *parsed_data['url'][point], headers=parsed_data['headers_value'][point], data=parsed_data['post_data'][point][verb])
                else:
                    r = requests.request(verb, *parsed_data['url'][point], headers=parsed_data['headers_value'][point])
                resp = r.json()
            except Exception as e:
                print(e)
                resp = r.text
            print('log1')
            
            worksheet.write( verb_counter, point_counter+y_data_counter+7, str(parsed_data['response_code'][point][verb])+str(r))

            # workbook,worksheet,resp_counter, y_resp_counter = write_to_excel(parsed_data['response'][point][verb],verbose=3, current_worksheet=worksheet, workbook=workbook,x0=verb_counter,y0=point_counter+y_data_counter+8)

            print(point,verb)
            # print('partsed', parsed_data['response'][point][verb])
            # print('fresh', resp)
            print('log2')
    
            # worksheet.write( verb_counter, point_counter+ y_data_counter + y_resp_counter+2+7, str(r) )
            # workbook,worksheet,resp_counter, y_resp_counter = write_to_excel(resp,verbose=3, current_worksheet=worksheet, workbook=workbook,x0=verb_counter,y0=point_counter+y_data_counter+y_resp_counter+2+8)
            print(parsed_data['response'][point][verb])
            try:
                resp2 = json.loads(parsed_data['response'][point][verb])
            except:
                resp2 = parsed_data['response'][point][verb]
                print()
                print()
                print(resp2)
                print()
                print()
            
            workbook,worksheet,resp_counter, y =  write_to_excel2( resp, resp2, shift=4, current_worksheet=worksheet, workbook=workbook,x0=verb_counter,y0=1+point_counter+y_data_counter+8)
            # workbook,worksheet,resp_counter, y =  write_to_excel2(resp, current_worksheet=worksheet, workbook=workbook,x0=verb_counter,y0=1+point_counter+8)
            # workbook,worksheet,resp_counter2, y2 =  write_to_excel2(parsed_data['response'][point][verb],  current_worksheet=worksheet, workbook=workbook,x0=verb_counter,y0=1+point_counter+6+8)
            print('resp',type(resp), len(resp))
            print("parsed_data['response'][point][verb]", type(parsed_data['response'][point][verb]), len( parsed_data['response'][point][verb]))
            print('log3')
            
            verb_counter += resp_counter+1
        worksheet.write( counter, point_counter+2, *parsed_data['url'][point])
        header_counter = counter
        for h in parsed_data['headers_name'][point]:
            worksheet.write( header_counter, point_counter+3, none_safe_str(h))
            worksheet.write( header_counter, point_counter+4, none_safe_str(parsed_data['headers_value'][point][h]))
            header_counter += 1
        counter = max(counter+1, verb_counter, header_counter) + 1
    workbook.close()
write_2_excel_parsed_data(parsed_data, settings )
    
    
    
    
    
    
write_response_data(parsed_data )




q=0     
for c in  x:
    q+=1
    if ifIter(c):
        print(type(c),len(c), q)
        if ifDict(c):
            for k0,v0 in c.items():
                print(k0)
                if q ==4:
                    printDict(v0)
                else:
                    print(v0)        
                    
        elif ifList(c):
            for v0 in c:
                if q ==4:
                    printDict(v0)
                else:
                    print(v0)        
        else:
            print(c)    
        print()
        # print(ifiter)
    else:
        print(type(c))
        print(c)
    print()

log

import os
cwd = os.getcwd()
cwd0 = os.getcwd()
os.chdir(r"C:\Users\syalosovetskyi\Downloads\python\middleware_script")
os.listdir('.')
cwd0

def prompt(s):
    print(s)
    ret = input()
    return ret
        

def user_interface():
    
    print('Интерфейс для программы тестирования')
    print('''Режимы работы: 
    1. считывание данных из файла,
    2. вывод текущих точек,
    3. вывод подробной информации о конкретной точке
    4. вывод всей информации
    5. тестирование точки''')
    

def init():
    
    s = int(user_interface() )
    
    
    
    
    token = fetchToken()
    # quirks = {'mode': 'change', 'change': 'url', 'change_type': 1, 'start_chr':'0', 'step_chr':'1', 'stop_chr':'150' }
    quirks = {'mode': 'content', 'content': 'header','change_type': 1,'step':10, 'chr':'a', 'content_num': 100}
    verbs = ['get', 'head', 'options', 'patch','connect', 'delete', 'post', 'put', 'trace']
    points = ['token', 'settings', 'links', 'relPhone', 'countMain', 'countDpi', 'bonus', 'offer']
    log = [{}, {}]
    asyncio.run( _requests(verbs, points, token,  debug=2 ))
    _write_response_to_excel('test')
